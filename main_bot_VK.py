# main.py
# Требует: pip install vk_api pandas
import json
import logging
import time
import traceback
import threading
import uuid
import re
import os
import sqlite3
import glob
import vk_api
import gspread
from google.oauth2.service_account import Credentials
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
import vk_api.utils
import config
from collections import OrderedDict
from functools import lru_cache
VK_TOKEN = getattr(config, 'VK_TOKEN', None)
GROUP_ID = getattr(config, 'GROUP_ID', None)
DB_PATH = getattr(config, 'DB_PATH', 'hosting.db')
import pandas as pd 
MAX_MEMORY_PAYMENTS = 50000   # Максимум выплат в памяти (сервер)
MAX_USER_CACHE_SIZE = 20000   # Максимум пользователей в кэше (сервер)
MEMORY_CLEANUP_INTERVAL = 600  # Очистка памяти каждые 10 минут (сервер)

def total_payments_count():
    """Подсчитывает общее количество выплат в памяти."""
    with user_payments_lock:
        return sum(len(payments) for payments in user_payments.values())

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger(__name__)
vk_session = vk_api.VkApi(token=VK_TOKEN)
vk = vk_session.get_api()
longpoll = VkBotLongPoll(vk_session, int(GROUP_ID))
user_payments = OrderedDict()
user_last_opened_payment = OrderedDict()  # Хранит последнюю открытую выплату для каждого пользователя

# Thread safety для многопоточного доступа
user_payments_lock = threading.RLock()
_csv_cache = {}
_cache_timestamps = {}
GSHEET_ID = "16ieoQC7N1lnmdMuonO3c7qdn_zmydptFYvRGSCjeLFg"
REPET_GSHEET_ID = "1UQMNS3yhFNCDyXS2E03y9iZX2zsHsoL3KKATo-e5c5Q"  # Таблица для репетиторов
_gspread_client = None

def ensure_db_indexes():
    """Создает индексы для оптимизации запросов."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        c.execute("PRAGMA journal_mode=WAL")  # Write-Ahead Logging для лучшей производительности
        c.execute("PRAGMA synchronous=NORMAL")  # Баланс между скоростью и надежностью
        c.execute("PRAGMA cache_size=10000")   # Увеличиваем кэш БД
        c.execute("PRAGMA temp_store=MEMORY")  # Временные таблицы в памяти
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_vedomosti_users_vk_id ON vedomosti_users(vk_id)",
            "CREATE INDEX IF NOT EXISTS idx_vedomosti_users_state ON vedomosti_users(state)",
            "CREATE INDEX IF NOT EXISTS idx_vedomosti_users_filename ON vedomosti_users(original_filename)",
            "CREATE INDEX IF NOT EXISTS idx_vedomosti_users_archive_at ON vedomosti_users(archive_at)",
            "CREATE INDEX IF NOT EXISTS idx_vedomosti_users_status ON vedomosti_users(status)",
            "CREATE INDEX IF NOT EXISTS idx_vedomosti_users_created_at ON vedomosti_users(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_vedomosti_users_confirmed_at ON vedomosti_users(confirmed_at)"
        ]
        
        for index_sql in indexes:
            c.execute(index_sql)
        
        conn.commit()
        conn.close()
        log.info("Database indexes and optimizations created successfully")
    except Exception:
        log.exception("Failed to create database indexes")

def cleanup_memory():
    try:
        # Ограничиваем общее количество выплат
        total = total_payments_count()
        if total > MAX_MEMORY_PAYMENTS:
            to_remove = total - MAX_MEMORY_PAYMENTS
            log.info("Need to remove %d old payments to respect MAX_MEMORY_PAYMENTS", to_remove)
            # Удаляем старые выплаты по пользователям (FIFO)
            for uid in list(user_payments.keys()):
                while user_payments[uid] and to_remove > 0:
                    user_payments[uid].pop(0)  # Удаляем самую старую выплату
                    to_remove -= 1
                if not user_payments[uid]:
                    del user_payments[uid]
                if to_remove <= 0:
                    break
            log.info("Cleaned up old payments from memory")
        
        if len(user_last_opened_payment) > MAX_USER_CACHE_SIZE:
            excess = len(user_last_opened_payment) - MAX_USER_CACHE_SIZE
            for _ in range(excess):
                user_last_opened_payment.popitem(last=False)
            log.info("Cleaned up %d old user cache entries", excess)
        
        current_time = time.time()
        expired_keys = [k for k, v in _cache_timestamps.items() if current_time - v > 3600]
        for key in expired_keys:
            _csv_cache.pop(key, None)
            _cache_timestamps.pop(key, None)
        
        if expired_keys:
            log.info("Cleaned up %d expired CSV cache entries", len(expired_keys))
            
    except Exception:
        log.exception("Failed to cleanup memory")

def get_cached_csv_data(file_path: str, ttl: int = 300):
    """Кэширует данные CSV файлов для ускорения доступа с проверкой mtime."""
    try:
        if not os.path.exists(file_path):
            return None
        
        mtime = os.path.getmtime(file_path)
        cache_key = file_path
        current_time = time.time()
        
        # Проверяем кэш с учетом mtime файла
        entry = _csv_cache.get(cache_key)
        if entry:
            df, cached_mtime, cache_time = entry
            if cached_mtime == mtime and (current_time - cache_time) < ttl:
                return df
        
        # Читаем файл
        try:
            df = pd.read_csv(file_path, dtype=str)
        except Exception:
            df = pd.read_csv(file_path, encoding='cp1251', dtype=str)
        
        if isinstance(df, pd.DataFrame) and not df.empty:
            # Кэшируем данные с mtime
            _csv_cache[cache_key] = (df, mtime, current_time)
            _cache_timestamps[cache_key] = current_time
            return df
        
        return None
    except Exception:
        log.exception("Failed to cache CSV data for %s", file_path)
        return None

def safe_vk_send(user_id: int, message: str, keyboard=None, max_retries: int = 3, delay: float = 1.0):
    """Унифицированная функция отправки сообщений VK с retry и проверкой длины."""
    # Проверяем длину сообщения
    if len(message) > 3800:
        log.warning("Message too long (%d chars) for user %s, truncating", len(message), user_id)
        message = message[:3800] + "..."
    
    for attempt in range(max_retries):
        try:
            params = {
                "user_id": user_id,
                "random_id": vk_api.utils.get_random_id(),
                "message": message
            }
            if keyboard is not None:
                params['keyboard'] = keyboard
            
            vk.messages.send(**params)
            return True
        except vk_api.exceptions.VkApiError as e:
            if e.code == 6:  # Rate limit
                wait_time = delay * (2 ** attempt)  # Exponential backoff
                log.warning("Rate limit hit for user %s, waiting %s seconds", user_id, wait_time)
                time.sleep(wait_time)
                continue
            elif e.code in [7, 9]:  # Permission denied or flood control
                log.warning("Permission denied or flood control for user %s: %s", user_id, e)
                return False
            else:
                log.warning("VK API error for user %s: %s", user_id, e)
                if attempt == max_retries - 1:
                    return False
                time.sleep(delay)
        except Exception as e:
            log.exception("Exception sending message to user %s: %s", user_id, e)
            if attempt == max_retries - 1:
                return False
            time.sleep(delay)
    
    return False

def send_vk_message_with_retry(user_id: int, message: str, max_retries: int = 3, delay: float = 1.0):
    """Обратная совместимость - использует safe_vk_send."""
    return safe_vk_send(user_id, message, None, max_retries, delay)

def get_gspread_client():
    global _gspread_client
    if _gspread_client is not None:
        log.debug("Using cached gspread client")
        return _gspread_client
    try:
        log.info("Initializing gspread client")
        with open(os.path.join(os.path.dirname(__file__), 'isu_groups.json'), 'r', encoding='utf-8') as f:
            info = json.load(f)
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_info(info, scopes=scopes)
        _gspread_client = gspread.authorize(creds)
        log.info("Successfully initialized gspread client")
        return _gspread_client
    except Exception as e:
        log.exception("Failed to init gspread client: %s", str(e))
        return None

def log_complaint_to_sheet(vk_id: int, reason: str, filename: str = "", filepath: str = "", fio: str = ""):
    try:
        log.info("Attempting to log complaint: vk_id=%s reason=%s filename=%s", vk_id, reason, filename)
        client = get_gspread_client()
        if not client:
            log.error("Failed to get gspread client")
            return
        fio_val = fio or ""
        if not fio_val:
            try:
                pays = get_payments_for_user(vk_id)
                if pays:
                    fio_val = pays[0].get("data", {}).get("fio") or pays[0].get("data", {}).get("curator") or ""
                if not fio_val:
                    allp = get_all_payments_for_user_from_db(vk_id, limit=1)
                    if allp:
                        fio_val = allp[0].get("data", {}).get("fio") or allp[0].get("data", {}).get("curator") or ""
            except Exception:
                fio_val = fio_val or ""
        log.info("Got gspread client, opening sheet with ID=%s", GSHEET_ID)
        sh = client.open_by_key(GSHEET_ID)
        ws = sh.sheet1
        dialog_link = f"https://vk.com/gim{GROUP_ID}?sel={vk_id}"
        vk_link = f"https://vk.com/id{vk_id}"
        row_data = [
            time.strftime('%Y-%m-%d %H:%M:%S'),
            vk_link,
            fio_val,
            reason,
            filename,
            filepath,
            dialog_link,
        ]
        log.info("Appending row: %s", row_data)
        ws.append_row(row_data, value_input_option='RAW')
        log.info("Successfully logged complaint to sheet for vk_id=%s reason=%s", vk_id, reason)
    except Exception as e:
        log.exception("Failed to log complaint to sheet for vk_id=%s reason=%s error=%s", vk_id, reason, str(e))


def log_repet_complaint_to_sheet(vk_id: int, reason: str, filename: str = "", fio: str = ""):
    """Записывает жалобу репетитора в Google таблицу."""
    try:
        log.info("Attempting to log repet complaint: vk_id=%s reason=%s filename=%s fio=%s", vk_id, reason, filename, fio)
        client = get_gspread_client()
        if not client:
            log.error("Failed to get gspread client")
            return
        log.info("Got gspread client, opening repet sheet with ID=%s", REPET_GSHEET_ID)
        sh = client.open_by_key(REPET_GSHEET_ID)
        ws = sh.sheet1
        dialog_link = f"https://vk.com/gim{GROUP_ID}?sel={vk_id}"
        # Столбцы: Дата, vk куратора, ФИО куратора, Причина несогласия, Название ведомости, Ссылка на диалог
        row_data = [time.strftime('%Y-%m-%d %H:%M:%S'), str(vk_id), fio, reason, filename, dialog_link]
        log.info("Appending repet row: %s", row_data)
        ws.append_row(row_data, value_input_option='RAW')
        log.info("Successfully logged repet complaint to sheet for vk_id=%s reason=%s", vk_id, reason)
    except Exception as e:
        log.exception("Failed to log repet complaint to sheet for vk_id=%s reason=%s error=%s", vk_id, reason, str(e))


def ensure_vedomosti_status_columns():
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        c.execute("PRAGMA table_info(vedomosti_users)")
        cols = [r[1] for r in c.fetchall()]
        if 'status' not in cols:
            c.execute("ALTER TABLE vedomosti_users ADD COLUMN status TEXT DEFAULT ''")
        if 'disagree_reason' not in cols:
            c.execute("ALTER TABLE vedomosti_users ADD COLUMN disagree_reason TEXT DEFAULT ''")
        if 'confirmed_at' not in cols:
            c.execute("ALTER TABLE vedomosti_users ADD COLUMN confirmed_at INTEGER")
        conn.commit()
        conn.close()
        log.info("ensure_vedomosti_status_columns done")
    except Exception:
        log.exception("Failed to ensure vedomosti status columns")


def ensure_unique_import_states():
    """Миграция: заменяет общие состояния вида 'imported:<suffix>' на уникальные UUID.
    Если суффикс не является UUID (например, это имя файла), генерируем новый UUID на КАЖДУЮ запись.
    Это предотвращает ситуацию, когда одно состояние разделяется несколькими пользователями.
    """
    try:
        if not os.path.exists(DB_PATH):
            return
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        c.execute("PRAGMA journal_mode=WAL")
        c.execute("SELECT id, state FROM vedomosti_users WHERE state LIKE 'imported:%'")
        rows = c.fetchall()
        updated = 0
        for db_id, state_val in rows:
            try:
                if not state_val:
                    continue
                parts = str(state_val).split(':', 1)
                if len(parts) != 2:
                    continue
                suffix = parts[1]
                try:
                    uuid.UUID(str(suffix))
                    is_uuid = True
                except Exception:
                    is_uuid = False
                if not is_uuid:
                    new_state = f"imported:{uuid.uuid4()}"
                    c.execute("UPDATE vedomosti_users SET state = ? WHERE id = ?", (new_state, db_id))
                    updated += 1
            except Exception:
                log.debug("Failed to migrate state for id=%s", db_id, exc_info=True)
        if updated:
            conn.commit()
            log.info("Migrated %d vedomosti states to unique UUID-based values", updated)
        conn.close()
    except Exception:
        log.exception("Failed to ensure unique import states")


def update_vedomosti_status_by_payment(payment_id: str, status: str, reason: str = None):
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        now = int(time.time())
        
        # Обрабатываем новый формат unique_payment_id (original_payment_id_db_id)
        if '_' in payment_id:
            try:
                db_id = int(payment_id.split('_')[-1])
                # Ищем по db_id (более точно)
                c.execute("SELECT id, vk_id, original_filename, state FROM vedomosti_users WHERE id = ?", (db_id,))
                existing_record = c.fetchone()
                
                if existing_record:
                    log.info("Found record by db_id for payment_id=%s: db_id=%s vk_id=%s filename=%s state=%s", 
                            payment_id, existing_record[0], existing_record[1], existing_record[2], existing_record[3])
                    
                    # Обновляем по db_id
                    if reason is not None:
                        c.execute(
                            "UPDATE vedomosti_users SET status = ?, disagree_reason = ?, confirmed_at = ? WHERE id = ?",
                            (status, str(reason), now, db_id)
                        )
                    else:
                        c.execute(
                            "UPDATE vedomosti_users SET status = ?, confirmed_at = ? WHERE id = ?",
                            (status, now, db_id)
                        )
                    conn.commit()
                    affected = c.rowcount
                    conn.close()
                    log.info("Updated vedomosti_users by db_id for payment=%s -> status=%s reason=%s affected=%s", payment_id, status, reason, affected)
                    
                    # Обновляем и в памяти
                    update_payment_in_memory(payment_id, status, reason)
                    return
                    
            except (ValueError, IndexError):
                # Если не удалось извлечь db_id, используем старый способ
                pass
        
        # Старый способ поиска по state (fallback)
        original_payment_id = payment_id.split('_')[0] if '_' in payment_id else payment_id
        state_val = f"imported:{original_payment_id}"
        
        c.execute("SELECT id, vk_id, original_filename FROM vedomosti_users WHERE state = ?", (state_val,))
        existing_record = c.fetchone()
        
        if not existing_record:
            log.error("No record found for payment_id=%s with state=%s", payment_id, state_val)
            conn.close()
            return
        
        log.info("Found record for payment_id=%s: id=%s vk_id=%s filename=%s", 
                payment_id, existing_record[0], existing_record[1], existing_record[2])
        
        if reason is not None:
            c.execute(
                "UPDATE vedomosti_users SET status = ?, disagree_reason = ?, confirmed_at = ? WHERE state = ?",
                (status, str(reason), now, state_val)
            )
        else:
            c.execute(
                "UPDATE vedomosti_users SET status = ?, confirmed_at = ? WHERE state = ?",
                (status, now, state_val)
            )
        conn.commit()
        affected = c.rowcount
        conn.close()
        log.info("Updated vedomosti_users for payment=%s -> status=%s reason=%s affected=%s", payment_id, status, reason, affected)
        
        # Обновляем в памяти
        update_payment_in_memory(payment_id, status, reason)
        
    except Exception:
        log.exception("Failed to update vedomosti status for payment %s", payment_id)


def update_payment_in_memory(payment_id: str, status: str, reason: str = None):
    """Обновляет статус платежа в памяти."""
    try:
        updated_count = 0
        for user_id, payments in user_payments.items():
            for payment in payments:
                # Проверяем и исходный payment_id и уникальный
                if payment["id"] == payment_id or payment.get("original_payment_id") == payment_id:
                    old_status = payment.get("status", "unknown")
                    payment["status"] = status
                    if reason is not None:
                        payment["disagree_reason"] = reason
                    log.info("Updated payment %s status %s->%s in memory for user %s (db_id=%s)", 
                            payment_id, old_status, status, user_id, payment.get("db_id"))
                    updated_count += 1
                    break
        if updated_count == 0:
            log.warning("No payments found in memory to update for payment_id %s", payment_id)
    except Exception:
        log.exception("Failed to update payment in memory for payment_id %s", payment_id)

def fetch_unprocessed_vedomosti():
    rows = []
    if not os.path.exists(DB_PATH):
        log.warning("DB file not found: %s", DB_PATH)
        return rows
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        c.execute("SELECT id, vk_id, personal_path, original_filename, state FROM vedomosti_users WHERE state IS NULL OR state = ''")
        rows = c.fetchall()
        conn.close()
    except Exception:
        log.exception("Failed to fetch vedomosti from sqlite")
    return rows


def mark_vedomosti_state(db_id, new_state):
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        c.execute("UPDATE vedomosti_users SET state = ? WHERE id = ?", (new_state, db_id))
        conn.commit()
        conn.close()
    except Exception:
        log.exception("Failed to update vedomosti state for id=%s", db_id)

def _map_row_to_payment_data(row_dict, vk_id, original_filename):
    def pick(*keys):
        for k in keys:
            if k is None:
                continue
            if k in row_dict and pd.notna(row_dict[k]):
                return str(row_dict[k])
        return ''
    data = {}
    data['fio'] = pick('ФИО', 'fio', 'name', 'full_name', 'FIO')
    data['phone'] = pick('Телефон', 'phone', 'Phone', 'telephone')
    data['console'] = pick('console', 'Console')
    data['curator'] = pick('Куратор', 'curator', 'manager', 'curator_name')
    data['vk_id'] = str(vk_id)
    data['mail'] = pick('Почта', 'mail', 'email', 'Email')
    data['groups'] = pick('Группы', 'groups', 'group', 'groups_list')
    for k in ['total_children','with_tutor','salary_per_student','salary_sum','retention','retention_pay',
              'okk','okk_pay','kpi_sum','checks_calc','checks_sum','extra_checks','support','webinars',
              'chats','group_calls','individual_calls','orders_table','bonus','penalties','total']:
        data[k] = pick(k, k.capitalize(), k.upper(), k.replace('_',' ').capitalize())
    data['total'] = data.get('total') or pick('Итого', 'Total', 'total')
    data['original_filename'] = os.path.basename(original_filename) if original_filename else ''
    return data


def _map_row_to_repet_payment_data(row_dict, vk_id, original_filename):
    """Маппинг данных для репетиторов из нового шаблона."""
    def pick(*keys):
        for k in keys:
            if k is None:
                continue
            if k in row_dict and pd.notna(row_dict[k]):
                return str(row_dict[k])
        return ''
    data = {}
    data['fio'] = pick('Репетитор', 'ФИО', 'fio', 'name', 'full_name', 'FIO')
    data['phone'] = pick('Номер', 'Телефон', 'phone', 'Phone', 'telephone')
    data['console'] = pick('console', 'Console')
    data['curator'] = pick('Репетитор', 'Куратор', 'curator', 'manager', 'curator_name')
    data['vk_id'] = str(vk_id)
    data['mail'] = pick('Почта', 'mail', 'email', 'Email')
    data['groups'] = pick('Группы', 'groups', 'group', 'groups_list')
    # Маппинг полей для репетиторов
    data['subject'] = pick('Предмет')
    data['lessons_held'] = pick('Кол-во состоявшихся занятий')
    data['lessons_no_student'] = pick('Кол-во занятий, на которые не явился ученик')
    data['base_payment'] = pick('Базовое вознаграждение за проведенные занятия')
    data['okk'] = pick('OKK', 'ОКК')
    data['rr'] = pick('RR')
    data['kpi'] = pick('KPI')
    data['preparation'] = pick('Подготовка к занятиям')
    data['penalties'] = pick('Штраф')
    data['total'] = pick('ИТОГ', 'Итого', 'Total', 'total')
    data['original_filename'] = os.path.basename(original_filename) if original_filename else ''
    return data


def import_vedomosti_into_memory(send_immediately: bool = False, rate_limit_delay: float = 0.35):
    rows = fetch_unprocessed_vedomosti()
    if not rows:
        return 0
    processed = 0
    for db_row in rows:
        try:
            db_id, vk_id_raw, personal_path, original_filename, state = db_row
            vk_id_raw = (vk_id_raw or '').strip()
            if not vk_id_raw:
                mark_vedomosti_state(db_id, 'no_vk')
                log.info("vedomosti id=%s has empty vk_id -> marked no_vk", db_id)
                continue
            try:
                vk_uid = int(vk_id_raw)
            except Exception:
                m = re.search(r'(\d{5,})', vk_id_raw)
                if m:
                    vk_uid = int(m.group(1))
                else:
                    mark_vedomosti_state(db_id, 'invalid_vk')
                    log.info("vedomosti id=%s has invalid vk_id=%s -> marked invalid_vk", db_id, vk_id_raw)
                    continue
            row_dict = {}
            if personal_path and os.path.exists(personal_path):
                try:
                    df = get_cached_csv_data(personal_path)
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        row_dict = df.iloc[0].to_dict()
                    else:
                        row_dict = {}
                except Exception:
                    log.warning("personal_path not found or empty for id=%s path=%s", db_id, personal_path)
            else:
                log.warning("personal_path not found or empty for id=%s path=%s", db_id, personal_path)
            
            # Проверяем, является ли это выплатой репетитора (по наличию столбца "Репетитор" или "Номер" или по state)
            is_repet = False
            if state and state.startswith('repet_imported:'):
                is_repet = True
            elif isinstance(row_dict, dict):
                is_repet = 'Репетитор' in row_dict or 'Номер' in row_dict
            
            if is_repet:
                payment_data = _map_row_to_repet_payment_data(row_dict, vk_uid, original_filename)
            else:
                payment_data = _map_row_to_payment_data(row_dict, vk_uid, original_filename)

            # Пропускаем записи с total == 0 (не записываем в память и не отправляем уведомления)
            def _is_zero_total(val) -> bool:
                if val is None:
                    return True
                s = str(val).strip()
                if s == "":
                    return True
                # заменяем запятую для float
                s_norm = s.replace(",", ".")
                try:
                    num = float(s_norm)
                    return abs(num) < 1e-9
                except Exception:
                    return False

            if _is_zero_total(payment_data.get('total')):
                mark_vedomosti_state(db_id, 'skip_zero_total')
                log.info("Skipped vedomosti id=%s for vk=%s (total=0)", db_id, vk_uid)
                continue

            # Сохраняем информацию о типе выплаты в данных
            if is_repet:
                payment_data['is_repet'] = True
            else:
                payment_data['is_repet'] = False

            pid = add_payment_for_user(vk_uid, payment_data)
            
            if is_repet:
                mark_vedomosti_state(db_id, f"repet_imported:{pid}")
            else:
                mark_vedomosti_state(db_id, f"imported:{pid}")
            
            processed += 1
            log.info("Imported vedomosti id=%s -> payment %s for vk=%s (file=%s, is_repet=%s)", db_id, pid, vk_uid, original_filename, is_repet)
            if send_immediately:
                try:
                    send_payment_message(vk_uid, find_payment(vk_uid, pid))
                    time.sleep(rate_limit_delay)
                except Exception:
                    log.exception("Failed to send immediate VK notification for payment %s vk=%s", pid, vk_uid)
        except Exception:
            log.exception("Failed to import vedomosti row %s", db_row)
    return processed


def background_importer(poll_interval=5.0):
    log.info("Background DB importer started (DB_PATH=%s, interval=%.1fs)", DB_PATH, poll_interval)
    last_cleanup = time.time()
    
    while True:
        try:
            # Import new reports and send notifications immediately to VK users
            imported = import_vedomosti_into_memory(send_immediately=True)
            if imported:
                log.info("Imported %d vedomosti into in-memory payments and sent VK notifications", imported)
            cleanup_archived_payments()
            current_time = time.time()
            if current_time - last_cleanup > MEMORY_CLEANUP_INTERVAL:
                cleanup_memory()
                last_cleanup = current_time
                
        except Exception:
            log.exception("Background importer exception")
        time.sleep(poll_interval)

def load_imported_vedomosti_into_memory(send_notifications: bool = False, rate_limit_delay: float = 0.35) -> int:
    if not os.path.exists(DB_PATH):
        log.warning("DB file not found: %s", DB_PATH)
        return 0
    loaded = 0
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        c.execute("SELECT id, vk_id, personal_path, original_filename, state, status, disagree_reason, confirmed_at FROM vedomosti_users WHERE state LIKE 'imported:%' OR state LIKE 'repet_imported:%'")
        rows = c.fetchall()
        conn.close()
    except Exception:
        log.exception("Failed to query imported vedomosti from sqlite")
        return 0

    for db_row in rows:
        try:
            db_id, vk_id_raw, personal_path, original_filename, state, status_db, disagree_reason_db, confirmed_at_db = db_row
            if not state or (not state.startswith('imported:') and not state.startswith('repet_imported:')):
                continue
            
            is_repet = state.startswith('repet_imported:')
            prefix = 'repet_imported:' if is_repet else 'imported:'
            parts = state.split(prefix, 1)
            if len(parts) != 2 or not parts[1]:
                continue
            payment_id = parts[1]
            vk_id_raw = (vk_id_raw or '').strip()
            if not vk_id_raw:
                log.info("Skipping imported row %s because vk_id empty", db_id)
                continue
            try:
                vk_uid = int(vk_id_raw)
            except Exception:
                m = re.search(r'(\d{5,})', vk_id_raw)
                if m:
                    vk_uid = int(m.group(1))
                else:
                    log.info("Skipping imported row %s due invalid vk_id=%s", db_id, vk_id_raw)
                    continue
            if find_payment(vk_uid, payment_id):
                log.debug("Payment %s already loaded for user %s, skipping", payment_id, vk_uid)
                continue
            row_dict = {}
            if personal_path and os.path.exists(personal_path):
                try:
                    df = get_cached_csv_data(personal_path)
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        row_dict = df.iloc[0].to_dict()
                except Exception:
                    pass
            
            if is_repet:
                payment_data = _map_row_to_repet_payment_data(row_dict, vk_uid, original_filename)
                payment_data['is_repet'] = True
            else:
                payment_data = _map_row_to_payment_data(row_dict, vk_uid, original_filename)
                payment_data['is_repet'] = False
            entry = {
                "id": payment_id,
                "data": payment_data,
                "created_at": float(confirmed_at_db) if confirmed_at_db else time.time(),
                "status": status_db or "new",
            }
            if disagree_reason_db:
                entry["disagree_reason"] = disagree_reason_db
            user_payments.setdefault(vk_uid, []).append(entry)
            loaded += 1
            log.info("Loaded imported vedomosti db_id=%s -> payment %s for vk=%s (file=%s) status=%s is_repet=%s", db_id, payment_id, vk_uid, original_filename, status_db, is_repet)
            if send_notifications:
                try:
                    send_payment_message(vk_uid, find_payment(vk_uid, payment_id))
                    time.sleep(rate_limit_delay)
                except Exception:
                    log.exception("Failed to send startup notification for payment %s vk=%s", payment_id, vk_uid)
        except Exception:
            log.exception("Error while loading imported vedomosti row %s", db_row)

    log.info("Loaded %d imported vedomosti into memory", loaded)
    return loaded

def cleanup_archived_payments():
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        c.execute("SELECT DISTINCT original_filename FROM vedomosti_users WHERE state LIKE 'imported:%' OR state LIKE 'repet_imported:%'")
        active_files = {row[0] for row in c.fetchall()}
        conn.close()
        
        removed_count = 0
        for user_id, payments in list(user_payments.items()):
            original_payments = payments.copy()
            payments[:] = [p for p in original_payments if p["data"].get("original_filename") in active_files]
            removed_count += len(original_payments) - len(payments)
            
            # Если у пользователя не осталось выплат, удаляем его из словаря
            if not payments:
                del user_payments[user_id]
        
        if removed_count > 0:
            log.info("Cleaned up %d archived payments from memory", removed_count)
        
    except Exception:
        log.exception("Failed to cleanup archived payments from memory")

def inline_confirm_keyboard(payment_id: str):
    kb = {
        "inline": True,
        "buttons": [
            [
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": "confirm_payment", "payment_id": payment_id, "choice": "agree"}, ensure_ascii=False),
                        "label": "Согласен с выплатой"
                    },
                    "color": "positive"
                },
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": "confirm_payment", "payment_id": payment_id, "choice": "disagree"}, ensure_ascii=False),
                        "label": "Не согласен с выплатой"
                    },
                    "color": "negative"
                }
            ]
        ]
    }
    return json.dumps(kb, ensure_ascii=False)


def yes_no_keyboard(cmd: str, payment_id: str):
    kb = {
        "inline": True,
        "buttons": [
            [
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": cmd, "payment_id": payment_id, "choice": "yes"}, ensure_ascii=False),
                        "label": "ДА"
                    },
                    "color": "positive"
                },
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": cmd, "payment_id": payment_id, "choice": "no"}, ensure_ascii=False),
                        "label": "НЕТ"
                    },
                    "color": "negative"
                }
            ]
        ]
    }
    return json.dumps(kb, ensure_ascii=False)

def final_agreement_keyboard(payment_id: str):
    kb = {
        "inline": True,
        "buttons": [
            [
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": "final_agreement", "payment_id": payment_id, "choice": "yes"}, ensure_ascii=False),
                        "label": "ДА"
                    },
                    "color": "positive"
                },
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": "final_agreement", "payment_id": payment_id, "choice": "no"}, ensure_ascii=False),
                        "label": "НЕТ"
                    },
                    "color": "negative"
                }
            ]
        ]
    }
    return json.dumps(kb, ensure_ascii=False)

def chat_bottom_keyboard():
    kb = {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": "to_list"}, ensure_ascii=False),
                        "label": "К списку выплат"
                    },
                    "color": "primary"
                }
            ]
        ]
    }
    return json.dumps(kb, ensure_ascii=False)


def payments_list_keyboard(statements, page: int = 0, page_size: int = 6):
    total = len(statements)
    start = page * page_size
    end = start + page_size
    page_items = statements[start:end]
    rows = []
    for sid, label in page_items:
        # Ограничиваем длину кнопки для соответствия лимитам VK
        if len(label) > 40:
            button_label = label[:37] + "..."
        else:
            button_label = label
            
        button_color = "primary"
        rows.append([
            {
                "action": {
                    "type": "text",
                    "payload": json.dumps({"cmd": "open_statement", "statement_id": sid}, ensure_ascii=False),
                    "label": button_label
                },
                "color": button_color
            }
        ])
    if end < total:
        next_page = page + 1
        rows.append([
            {
                "action": {
                    "type": "text",
                    "payload": json.dumps({"cmd": "payments_page", "page": next_page}, ensure_ascii=False),
                    "label": "Ещё"
                },
                "color": "secondary"
            }
        ])
    kb = {"inline": True, "buttons": rows}
    return json.dumps(kb, ensure_ascii=False)


def payments_list_keyboard_for_user(user_payments_list, page: int = 0, page_size: int = 5, add_more: bool = True):
    """Клавиатура списка выплат для конкретного пользователя с учетом его статусов."""
    total = len(user_payments_list)
    start = page * page_size
    end = start + page_size
    page_items = user_payments_list[start:end]
    rows = []
    for idx, entry in enumerate(page_items, start=1 + start):
        sid = entry.get("id")
        status = entry.get("status")
        
        # Учитываем статус при расчете максимальной длины
        if status == "agreed":
            # Оставляем место для " " (3 символа) 
            max_label_length = 37
            base_label = _format_payment_label(
                entry.get("data", {}).get('original_filename'), 
                idx, 
                max_label_length,
                entry.get("created_at"),
                entry.get("db_id"),
                entry.get("data", {}).get("groups")  # Передаем информацию о группах
            )
            button_label = f"{base_label} "
            button_color = "positive"
        else:
            # Полная длина для обычных кнопок
            base_label = _format_payment_label(
                entry.get("data", {}).get('original_filename'), 
                idx, 
                40,
                entry.get("created_at"),
                entry.get("db_id"),
                entry.get("data", {}).get("groups")  # Передаем информацию о группах
            )
            button_label = base_label
            button_color = "primary"
            
        rows.append([
            {
                "action": {
                    "type": "text",
                    "payload": json.dumps({"cmd": "open_statement", "statement_id": sid}, ensure_ascii=False),
                    "label": button_label
                },
                "color": button_color
            }
        ])
    if add_more and end < total:
        next_page = page + 1
        rows.append([
            {
                "action": {
                    "type": "text",
                    "payload": json.dumps({"cmd": "payments_page", "page": next_page}, ensure_ascii=False),
                    "label": "Ещё"
                },
                "color": "secondary"
            }
        ])
    kb = {"inline": True, "buttons": rows}
    return json.dumps(kb, ensure_ascii=False)

def send_payments_list_multiple(user_id: int, payments: list, page: int = 0, use_vk_direct: bool = False):
    """Отправляет список выплат, разбивая на несколько сообщений если кнопок больше 5.
    
    Args:
        user_id: ID пользователя VK
        payments: Список выплат
        page: Номер страницы (0 для первой)
        use_vk_direct: Если True, использует vk.messages.send напрямую, иначе safe_vk_send
    """
    if not payments:
        if use_vk_direct:
            vk.messages.send(user_id=user_id, random_id=vk_api.utils.get_random_id(), 
                           message="У Вас нет выплат.", keyboard=chat_bottom_keyboard())
        else:
            safe_vk_send(user_id, "У Вас нет выплат.", chat_bottom_keyboard())
        return
    
    total = len(payments)
    page_size = 5  # Безопасный лимит для VK API
    
    # Если выплат <= 5, отправляем одним сообщением
    if total <= page_size:
        if page == 0:
            message = "Список ведомостей (выберите):"
        else:
            message = f"Список ведомостей (страница {page+1}):"
        keyboard = payments_list_keyboard_for_user(payments, page=page, page_size=page_size)
        if use_vk_direct:
            vk.messages.send(user_id=user_id, random_id=vk_api.utils.get_random_id(), 
                           message=message, keyboard=keyboard)
        else:
            safe_vk_send(user_id, message, keyboard)
        return
    
    # Если выплат больше 5 и это первая страница, отправляем несколько сообщений сразу
    if page == 0:
        # Отправляем первое сообщение с первыми 5 кнопками, без кнопки "Ещё"
        keyboard = payments_list_keyboard_for_user(payments, page=0, page_size=page_size, add_more=False)
        if use_vk_direct:
            vk.messages.send(user_id=user_id, random_id=vk_api.utils.get_random_id(), 
                           message="Список ведомостей (выберите):", keyboard=keyboard)
        else:
            safe_vk_send(user_id, "Список ведомостей (выберите):", keyboard)
        
        # Отправляем оставшиеся части
        remaining_payments = payments[page_size:]
        part_num = 2
        
        for i in range(0, len(remaining_payments), page_size):
            part_payments = remaining_payments[i:i+page_size]
            start_idx = page_size + i + 1  # Начальный индекс для нумерации
            
            # Создаем клавиатуру для этой части
            rows = []
            for idx, entry in enumerate(part_payments, start=start_idx):
                sid = entry.get("id")
                status = entry.get("status")
                
                if status == "agreed":
                    max_label_length = 37
                    base_label = _format_payment_label(
                        entry.get("data", {}).get('original_filename'), 
                        idx, 
                        max_label_length,
                        entry.get("created_at"),
                        entry.get("db_id"),
                        entry.get("data", {}).get("groups")
                    )
                    button_label = f"{base_label} "
                    button_color = "positive"
                else:
                    base_label = _format_payment_label(
                        entry.get("data", {}).get('original_filename'), 
                        idx, 
                        40,
                        entry.get("created_at"),
                        entry.get("db_id"),
                        entry.get("data", {}).get("groups")
                    )
                    button_label = base_label
                    button_color = "primary"
                
                rows.append([
                    {
                        "action": {
                            "type": "text",
                            "payload": json.dumps({"cmd": "open_statement", "statement_id": sid}, ensure_ascii=False),
                            "label": button_label
                        },
                        "color": button_color
                    }
                ])
            
            kb = {"inline": True, "buttons": rows}
            keyboard_json = json.dumps(kb, ensure_ascii=False)
            
            part_message = f"Список ведомостей (часть {part_num}):"
            if use_vk_direct:
                vk.messages.send(user_id=user_id, random_id=vk_api.utils.get_random_id(), 
                               message=part_message, keyboard=keyboard_json)
            else:
                safe_vk_send(user_id, part_message, keyboard_json)
            part_num += 1
    else:
        # Для страниц > 0 используем обычную пагинацию
        message = f"Список ведомостей (страница {page+1}):"
        keyboard = payments_list_keyboard_for_user(payments, page=page, page_size=page_size)
        if use_vk_direct:
            vk.messages.send(user_id=user_id, random_id=vk_api.utils.get_random_id(), 
                           message=message, keyboard=keyboard)
        else:
            safe_vk_send(user_id, message, keyboard)


def payments_disagree_keyboard(payment_id: str):
    labels = [
        "Число учеников",
        "Проверки ДЗ",
        "Штрафы",
        "Стол заказов",
        "Вебинары",
        "Оплата за УП",
        "Оплата за чаты",
        "КПИ за продления",
        "Иная причина (связаться с оператором)"
    ]
    if len(labels) == 0:
        return json.dumps({"inline": True, "buttons": []}, ensure_ascii=False)
    last_label = labels[-1]
    first_labels = labels[:-1]
    rows = []
    for i in range(0, len(first_labels), 2):
        chunk = first_labels[i:i+2]
        row = []
        for lab in chunk:
            btn = {
                "action": {
                    "type": "text",
                    "payload": json.dumps({"cmd": "disagree_reason", "payment_id": payment_id, "reason": lab}, ensure_ascii=False),
                    "label": lab
                },
                "color": "primary"
            }
            row.append(btn)
        rows.append(row)
    last_btn = {
        "action": {
            "type": "text",
            "payload": json.dumps({"cmd": "disagree_reason", "payment_id": payment_id, "reason": last_label}, ensure_ascii=False),
            "label": last_label
        },
        "color": "primary"
    }
    rows.append([last_btn])
    
    # Добавляем кнопку "Согласиться с ведомостью"
    agree_btn = {
        "action": {
            "type": "text",
            "payload": json.dumps({"cmd": "agree_payment", "payment_id": payment_id}, ensure_ascii=False),
            "label": "Согласиться с ведомостью"
        },
        "color": "positive"
    }
    rows.append([agree_btn])
    
    kb = {"inline": True, "buttons": rows}
    return json.dumps(kb, ensure_ascii=False)
def disagreement_decision_keyboard(payment_id: str, reason_label: str):
    kb = {
        "inline": True,
        "buttons": [
            [
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": "disagree_decision", "payment_id": payment_id, "reason": reason_label, "choice": "agree_point"}, ensure_ascii=False),
                        "label": "Согласен с пунктом"
                    },
                    "color": "positive"
                },
                {
                    "action": {
                        "type": "text",
                        "payload": json.dumps({"cmd": "disagree_decision", "payment_id": payment_id, "reason": reason_label, "choice": "disagree_point"}, ensure_ascii=False),
                        "label": "Не согласен с пунктом"
                    },
                    "color": "negative"
                }
            ]
        ]
    }
    return json.dumps(kb, ensure_ascii=False)
def format_payment_text(data: dict) -> str:
    """Форматирует текст выплаты в красивом виде, используя данные из CSV файла."""
    try:
        vk_id_str = str(data.get('vk_id','')).strip()
        original_filename = data.get('original_filename') or ''
        
        # ИСПРАВЛЕНИЕ: Используем конкретный путь к файлу, если он есть в данных
        personal_path = data.get('personal_path')
        if personal_path and os.path.exists(personal_path):
            # Загружаем данные из конкретного персонального файла
            csv_path = personal_path
            log.debug("Using specific personal_path for format_payment_text: %s", csv_path)
        else:
            # Fallback на старый метод поиска (для совместимости)
            if not vk_id_str or not original_filename:
                return format_payment_text_fallback(data)
            
            # Получаем base_name для поиска CSV файла
            base_name = os.path.splitext(original_filename)[0] if original_filename else ''
            
            # Ищем CSV файл
            csv_path = _find_curator_csv(base_name, int(vk_id_str))
            if not csv_path:
                return format_payment_text_fallback(data)
        
        # Читаем CSV
        try:
            df = pd.read_csv(csv_path, dtype=str)
        except Exception:
            try:
                df = pd.read_csv(csv_path, encoding='cp1251', dtype=str)
            except Exception:
                return format_payment_text_fallback(data)
        
        if df is None or df.empty or 'vk_id' not in df.columns:
            return format_payment_text_fallback(data)
        
        # Находим строку пользователя
        vk_series = df['vk_id'].fillna('').astype(str).apply(_extract_numeric_vk)
        row = df[vk_series == vk_id_str]
        if row.empty:
            return format_payment_text_fallback(data)
        
        p = row.fillna('0').iloc[0]
        
        # Современный формат с учётом новых столбцов
        sections = _compose_payment_sections(p)
        checks_block = sections["checks"]
        if checks_block:
            checks_block = "\n[Проверки]" + checks_block
        extras_block = sections["extras"]
        if extras_block.startswith("\n[Иная деятельность]"):
            extras_block = "\n[ Иная деятельность]" + extras_block[len("\n[Иная деятельность]"):]
        fines_block = sections["fines"]
        if fines_block.startswith("\n\n→ Штрафы"):
            fines_block = "\n\n" + fines_block[len("\n\n→ "):]
        total_block = sections["total"]
        if total_block.startswith("\n\n→ ИТОГО"):
            total_block = "\n\n" + total_block[len("\n\n→ "):]
        
        groups_str = str(p.get('groups', '')).strip()
        groups_line = f"\nГруппы: {groups_str}" if groups_str else ""
        
        base = (f"=== Согласование выплаты ==="
                f"\nВедомость: {original_filename.replace('.csv', '').replace('_', ' ')}"
                f"\nКуратор: {p.get('name', '')}"
                f"\nТип куратора: {p.get('type', '')}"
                f"\nПочта на платформе: {p.get('email', '')}"
                f"{groups_line}\n")

        final = ("\n\nНажмите «Согласен», если у Вас нет разногласий с выставленными цифрами"
                 "\nНажмите «Не согласен», если Вы не согласны с каким-либо из пунктов"
                 "\nПросмотр ведомости возможен в течение 36 часов")

        msg = (base
               + sections["studs"]
               + sections["retention"]
               + sections["okk"]
               + checks_block
               + extras_block
               + fines_block
               + total_block
               + final)
        return msg
        
        # Используем красивое форматирование как в format_message
        base = (f"=== Согласование выплаты ==="
                f"\nВедомость: {original_filename.replace('.csv', '').replace('_', ' ')}"
                f"\nКуратор: {p.get('name', '')}"
                f"\nТип куратора: {p.get('type', '')}"
                f"\nПочта на платформе: {p.get('email', '')}\n")

        studs_section = ""
        stud_all = _to_int_safe(p.get('stud_all'))
        if stud_all > 0:
            studs_section = (f"\n[Сопровождение учеников]"
                             f"\nВсего учеников в группах: {stud_all}"
                             f"\nСтавка за ученика: {_to_int_safe(p.get('base'))}₽")
            stud_rep = _to_int_safe(p.get('stud_rep'))
            if stud_rep > 0:
                studs_section += (f"\nИз них с репетитором: {stud_rep}"
                                  f"\nДоплата за учеников с репетитором: 50₽ / чел")
            studs_section += f"\n→ Всего за сопровождение: {_to_int_safe(p.get('stud_salary'))}₽\n"
            studs_section += (f"\nПоказатель RR: {p.get('rr','')} | KPI за RR: {_to_float_str_money(p.get('rr_salary'))}₽"
                              f"\nПоказатель ОКК: {p.get('okk','')} | KPI за ОКК: {_to_float_str_money(p.get('okk_salary'))}₽"
                              f"\n→ Всего KPI (RR+OKK): {_to_float_str_money(p.get('kpi_total'))}₽\n")

        checks_section = ""
        checks_salary = _to_int_safe(p.get('checks_salary'))
        dop_checks = _to_int_safe(p.get('dop_checks'))
        if checks_salary > 0 or dop_checks > 0:
            checks_section = f"\n[Проверки]"
            if checks_salary > 0:
                checks_section += f"\n→ Проверка домашних работ: {checks_salary}₽"
            if dop_checks > 0:
                checks_section += f"\n→ Дополнительно – за проверки (данные СК): {dop_checks}₽"
            checks_section += "\n"

        # Дополнительная деятельность
        extras_keys = ['up','chats','webs','meth','dop_sk','callsg','callsp']
        extras_names = {
            'up': 'За учебную поддержку',
            'chats': 'Модерация чатов',
            'webs': 'Модерация вебинаров',
            'callsg': 'Групповые созвоны',
            'callsp': 'Индивидуальные созвоны',
            'dop_sk': 'Доп. суммы, начисленные СК',
            'meth': 'Стол заказов',
        }
        extras_total = sum(_to_int_safe(p.get(k)) for k in extras_keys)
        dops_section = ""
        if extras_total > 0:
            dops_section = "\n[ Иная деятельность]"
            for k in extras_keys:
                v = _to_int_safe(p.get(k))
                if v > 0:
                    dops_section += f"\n{extras_names[k]}: {v}₽"
            dops_section += f"\n→ Всего в категории: {extras_total}₽"

        # Штрафы
        fines_val = _to_int_safe(p.get('fines'))
        fines_section = f"\n\nШтрафы: -{fines_val}₽" if fines_val > 0 else f"\n\nШтрафы: отсутствуют"

        # Итого
        total_section = f"\n\nИТОГО К ВЫПЛАТЕ: {_to_float_str_money(p.get('total'))}₽"
        
        # Добавляем комментарий, если он осмысленный
        comment = p.get('comment', '')
        if _is_meaningful_comment(comment):
            total_section += f"\n[!!!] Комментарий: {str(comment).strip()}"
        
        # Финальная информация
        final = ("\n\nНажмите «Согласен», если у Вас нет разногласий с выставленными цифрами"
                 "\nНажмите «Не согласен», если Вы не согласны с каким-либо из пунктов"
                 "\nПросмотр ведомости возможен в течение 36 часов")

        msg = base + studs_section + checks_section + dops_section + fines_section + total_section + final
        return msg
        
    except Exception:
        log.exception("Failed to format payment text with enhanced format, using fallback")
        return format_payment_text_fallback(data)


def format_payment_text_fallback(data: dict) -> str:
    """Простое форматирование выплаты (fallback)."""
    lines = []
    lines.append("=== Согласование выплаты ===")
    lines.append(f"Номер телефона, который указан в консоли: {data.get('phone','')}")
    lines.append(f"ФИО, которое указано в консоли: {data.get('console','') or data.get('fio','')}")
    lines.append(f"Тип куратора: {data.get('type','')}")
    lines.append(f"Куратор: {data.get('curator','') or data.get('name','')}")
    lines.append(f"vk_id: {data.get('vk_id','')}")
    lines.append(f"Почта: {data.get('mail','')}")
    if _is_meaningful_comment(data.get('comment')):
        lines.append(f"Комментарий: {str(data.get('comment')).strip()}")
    lines.append(f"Группы: {data.get('groups','')}")

    sections = _compose_payment_sections(data)
    fallback_order = ["studs", "retention", "okk", "checks", "extras", "fines", "total"]
    for key in fallback_order:
        block = sections.get(key, "")
        if not block:
            continue
        if key == "checks":
            block = "\n[Проверки]" + block
        if key == "extras" and block.startswith("\n[Иная деятельность]"):
            block = "\n[ Иная деятельность]" + block[len("\n[Иная деятельность]"):]
        if key == "fines" and block.startswith("\n\n→ Штрафы"):
            block = "\n\n" + block[len("\n\n→ "):]
        if key == "total" and block.startswith("\n\n→ ИТОГО"):
            block = "\n\n" + block[len("\n\n→ "):]
        lines.append(block)

    lines.append("\nПросмотр ведомости возможен в течение 36 часов")
    return "\n".join(lines)


def format_repet_payment_text(data: dict) -> str:
    """Форматирует текст выплаты для репетиторов по новому шаблону."""
    try:
        vk_id_str = str(data.get('vk_id','')).strip()
        original_filename = data.get('original_filename') or ''
        
        # Используем конкретный путь к файлу, если он есть в данных
        personal_path = data.get('personal_path')
        if personal_path and os.path.exists(personal_path):
            csv_path = personal_path
            log.debug("Using specific personal_path for format_repet_payment_text: %s", csv_path)
        else:
            # Fallback на старый метод поиска
            if not vk_id_str or not original_filename:
                return format_repet_payment_text_fallback(data)
            
            base_name = os.path.splitext(original_filename)[0] if original_filename else ''
            csv_path = _find_curator_csv(base_name, int(vk_id_str))
            if not csv_path:
                return format_repet_payment_text_fallback(data)
        
        # Читаем CSV
        try:
            df = pd.read_csv(csv_path, dtype=str)
        except Exception:
            try:
                df = pd.read_csv(csv_path, encoding='cp1251', dtype=str)
            except Exception:
                return format_repet_payment_text_fallback(data)
        
        if df is None or df.empty:
            return format_repet_payment_text_fallback(data)
        
        # Берём первую строку (персональный файл должен содержать одну строку)
        p = df.fillna('0').iloc[0]
        
        # Форматируем сообщение по новому шаблону
        msg = "Открыта ведомость\n\n"
        msg += "=== Согласование выплаты ===\n"
        msg += f"ФИО: {p.get('Репетитор', '') or data.get('fio', '')}\n"
        msg += f"Предмет: {p.get('Предмет', '') or data.get('subject', '')}\n"
        msg += f"Кол-во проведенных занятий: {p.get('Кол-во состоявшихся занятий', '') or data.get('lessons_held', '')}\n"
        msg += f"Уроки без подключения ученика: {p.get('Кол-во занятий, на которые не явился ученик', '') or data.get('lessons_no_student', '')}\n"
        msg += f"Оплата за занятия: {p.get('Базовое вознаграждение за проведенные занятия', '') or data.get('base_payment', '')}\n"
        msg += f"Оценка контроля качества: {p.get('OKK', '') or p.get('ОКК', '') or data.get('okk', '')}\n"
        msg += f"Критерий удержания учеников: {p.get('RR', '') or data.get('rr', '')}\n"
        msg += f"Дополнительное вознаграждение за качество: {p.get('KPI', '') or data.get('kpi', '')}\n"
        # Приоритет: data['preparation'] (из маппинга), затем из CSV напрямую
        prep_val = data.get('preparation', '') or ''
        if not str(prep_val).strip() or str(prep_val).strip() == '0':
            prep_val = p.get('Подготовка к занятиям', '')
        prep_str = str(prep_val).strip() if prep_val else '0'
        msg += f"Доп. вознаграждение за подготовку: {prep_str}\n"
        msg += f"Штрафы: {p.get('Штраф', '') or data.get('penalties', '')}\n"
        msg += f"Итоговая сумма: {p.get('ИТОГ', '') or data.get('total', '')}\n\n"
        msg += "Нажмите «Согласен», если у Вас нет разногласий с выставленными цифрами\n"
        msg += "Нажмите «Не согласен», если Вы не согласны с каким-либо из пунктов\n"
        msg += "Просмотр ведомости возможен в течение 36 часов"
        
        return msg
        
    except Exception:
        log.exception("Failed to format repet payment text, using fallback")
        return format_repet_payment_text_fallback(data)


def format_repet_payment_text_fallback(data: dict) -> str:
    """Простое форматирование выплаты для репетиторов (fallback)."""
    lines = []
    lines.append("Открыта ведомость")
    lines.append("")
    lines.append("=== Согласование выплаты ===")
    lines.append(f"ФИО: {data.get('fio','')}")
    lines.append(f"Предмет: {data.get('subject','')}")
    lines.append(f"Кол-во проведенных занятий: {data.get('lessons_held','')}")
    lines.append(f"Уроки без подключения ученика: {data.get('lessons_no_student','')}")
    lines.append(f"Оплата за занятия: {data.get('base_payment','')}")
    lines.append(f"Оценка контроля качества: {data.get('okk','')}")
    lines.append(f"Критерий удержания учеников: {data.get('rr','')}")
    lines.append(f"Дополнительное вознаграждение за качество: {data.get('kpi','')}")
    lines.append(f"Доп. вознаграждение за подготовку: {data.get('preparation','')}")
    lines.append(f"Штрафы: {data.get('penalties','')}")
    lines.append(f"Итоговая сумма: {data.get('total','')}")
    lines.append("")
    lines.append("Нажмите «Согласен», если у Вас нет разногласий с выставленными цифрами")
    lines.append("Нажмите «Не согласен», если Вы не согласны с каким-либо из пунктов")
    lines.append("Просмотр ведомости возможен в течение 36 часов")
    return "\n".join(lines)


def map_reason_to_type(reason_label: str) -> str:
    mapping = {
        "Число учеников": "students",
        "Проверки ДЗ": "homework",
        "Штрафы": "fines",
        "Стол заказов": "meth",
        "Вебинары": "webs",
        "Оплата за УП": "up",
        "Оплата за чаты": "dops",
        "КПИ за продления": "rr",
    }
    return mapping.get(reason_label, "")

def format_conflict(file_name, uid, conflict_type, personal_path=None):
    if conflict_type == "students":
            reply = (f"Количество учеников взято из журнала оплат с последних продлений (листы «Статистика по группам», «Статистика по кураторам»). "
                     f"Результат просуммирован за все группы"
                     f"\n\nЕсли ученик записался на сразу 2-й блок и не занимался в 1-м, оплата за его сопровождение в 1-м блоке не последует")
    elif conflict_type == "homework":
            try:
                def find_homework_csv(base_name: str, vk_uid: int):
                    root = os.path.dirname(os.path.abspath(__file__))
                    patterns = [
                        os.path.join(root, "hosting", "open", "**", "users", f"{vk_uid}_*_{base_name}.csv"),
                        os.path.join(root, "hosting", "open", "**", "users", f"{vk_uid}_*_{base_name.replace(' ','_')}.csv"),
                    ]
                    matches = []
                    for pat in patterns:
                        matches.extend(glob.glob(pat, recursive=True))
                    if matches:
                        matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                        return matches[0]
                    group_patterns = [
                        os.path.join(root, "hosting", "open", "**", f"{base_name}.csv"),
                        os.path.join(root, "hosting", "open", "**", f"{base_name.replace(' ','_')}.csv"),
                    ]
                    group_matches = []
                    for pat in group_patterns:
                        group_matches.extend(glob.glob(pat, recursive=True))
                    if group_matches:
                        group_matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                        return group_matches[0]
                    return None

                csv_path = None
                if file_name:
                    csv_path = find_homework_csv(file_name, uid)
                if not csv_path:
                    raise FileNotFoundError("CSV for homework not found")
                try:
                    df = pd.read_csv(csv_path, dtype=str)
                except Exception:
                    df = pd.read_csv(csv_path, encoding='cp1251', dtype=str)
                if df is None or df.empty or 'vk_id' not in df.columns:
                    raise ValueError("CSV missing data or vk_id column")

                def _norm_vk(val):
                    s = str(val).strip()
                    m = re.search(r'(\d+)', s)
                    return m.group(1) if m else s

                df_valid = df[df['vk_id'].notna()].copy()
                df_valid['_vk_norm'] = df_valid['vk_id'].apply(_norm_vk)
                row = df_valid[df_valid['_vk_norm'].astype(str) == str(uid)]
                if row.empty:
                    raise ValueError("User row not found in CSV")
                p = row.fillna('0').iloc[0]

                def to_int_safe(v):
                    try:
                        return int(float(str(v)))
                    except Exception:
                        return 0

                abs_val = to_int_safe(p['checks_all'])
                prev_val = to_int_safe(p['checks_prev'])
                fin_val = to_int_safe(p['checks_salary'])
                reply = (f"Оплата за ДЗ считается как общая сумма за проверенные номера за всё время минус ранее оплаченные работы. Годовые и полугодовые курсы разделяются в вопросе расчета выплаты"
                         f"\n\nОбщая сумма твоих проверок за всё время на аккаунте: {abs_val}"
                         f"\nОбщая сумма твоих проверок на момент предыдущей выплаты: {prev_val}"
                         f"\nТаким образом, в эту выплату пойдёт: {abs_val} - {prev_val} = {fin_val}"
                         f"\n\nЕсли в какой-либо выгрузке ты видишь, что итоговая сумма уже больше, чем сейчас, то эта разница пойдет в следующую выплату")

            except Exception as e:
                log.exception("Error processing homework CSV: %s", e)
                reply = (f"Оплата за ДЗ считается как общая сумма за проверенные номера за всё время минус ранее оплаченные работы. Годовые и полугодовые курсы разделяются в вопросе расчета выплаты"
                         f"\n\nЕсли конкретные цифры недоступны, сверка по CSV будет выполнена оператором.")
    elif conflict_type == "fines":
            reply = (f"Штрафы выставляются старшими кураторами курса. Если ты не осведомлен(-а) о каком-либо вычете, уточни об этом у старшего куратора"
                     f"\nМы можем откорректировать сумму штрафа в выплате, если запрос на это передаст старший куратор")
    elif conflict_type == "meth":
            reply = (f"Оплата за стол заказов выставляется методистом предмета, по вопросам расчёта выплаты обращайся к нему"
                    f"\nМы можем откорректировать сумму за стол заказов в выплате, если запрос на это передаст методист")
    elif conflict_type == "webs":
            reply = (f"Сумму за вебы можно отследить в течение блока, тк вы самостоятельно заполняете отчетность по ним."
                     f"\nЕсли ты не заполнил(-а) все вебы за этот блок, то можешь их добавить в следующий. Данные за этот блок уже считаны, их не исправить")
    elif conflict_type == "up":
            reply = (f"Оплату УП выставляет руководитель УП, уточни, пожалуйста, у него этот момент"
                     f"\nМы можем откорректировать сумму за УП в выплате, если запрос на это передаст руководитель УП")
    elif conflict_type == "dops":
            reply = (f"Дополнительные выплаты (проверки, перепроверки) рассчитывает и выставляет старший куратор, уточни, пожалуйста, у него этот момент"
                     f"\nМы можем откорректировать сумму допов в выплате, если запрос на это передаст старший куратор")
    elif conflict_type == "rr":
        try:
            def find_rr_csv(base_name: str, vk_uid: int):
                root = os.path.dirname(os.path.abspath(__file__))
                patterns = [
                    os.path.join(root, "hosting", "open", "**", "users", f"{vk_uid}_*_{base_name}.csv"),
                    os.path.join(root, "hosting", "open", "**", "users", f"{vk_uid}_*_{base_name.replace(' ','_')}.csv"),
                ]
                matches = []
                for pat in patterns:
                    matches.extend(glob.glob(pat, recursive=True))
                if matches:
                    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                    return matches[0]
                group_patterns = [
                    os.path.join(root, "hosting", "open", "**", f"{base_name}.csv"),
                    os.path.join(root, "hosting", "open", "**", f"{base_name.replace(' ','_')}.csv"),
                ]
                group_matches = []
                for pat in group_patterns:
                    group_matches.extend(glob.glob(pat, recursive=True))
                if group_matches:
                    group_matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                    return group_matches[0]
                return None

            def find_excel_file(csv_path: str):
                """Находит исходный Excel файл по пути к CSV."""
                if not csv_path:
                    return None
                
                csv_dir = os.path.dirname(csv_path)
                csv_basename = os.path.basename(csv_path)
                csv_name_noext = os.path.splitext(csv_basename)[0]
                
                # Если CSV файл находится в папке users, ищем групповой CSV файл
                group_csv_path = None
                search_dir = csv_dir
                
                if 'users' in csv_dir:
                    parent_dir = os.path.dirname(csv_dir)
                    # Ищем групповой CSV файл в родительской директории
                    if os.path.exists(parent_dir):
                        for filename in os.listdir(parent_dir):
                            if filename.lower().endswith('.csv') and 'users' not in filename:
                                group_csv_path = os.path.join(parent_dir, filename)
                                search_dir = parent_dir  # Ищем Excel в той же директории, где групповой CSV
                                log.info("Found group CSV: %s, will search Excel in: %s", group_csv_path, search_dir)
                                break
                else:
                    group_csv_path = csv_path
                
                # Определяем базовое имя для поиска Excel (из группового CSV)
                if group_csv_path and os.path.exists(group_csv_path):
                    group_base = os.path.splitext(os.path.basename(group_csv_path))[0]
                else:
                    group_base = csv_name_noext
                
                # ПРИОРИТЕТ 1: Ищем Excel файл в той же директории, что и групповой CSV (hosting/open/...)
                if os.path.exists(search_dir):
                    for ext in ['.xlsx', '.xls']:
                        # Пробуем точное имя
                        excel_path = os.path.join(search_dir, group_base + ext)
                        if os.path.exists(excel_path):
                            log.info("Found Excel file in same directory as CSV: %s", excel_path)
                            return excel_path
                        
                        # Пробуем все Excel файлы в этой директории
                        for filename in os.listdir(search_dir):
                            if filename.lower().endswith(ext):
                                # Проверяем, совпадает ли базовое имя (с учетом замены пробелов/подчеркиваний)
                                file_base = os.path.splitext(filename)[0]
                                if (group_base in file_base or file_base in group_base or 
                                    group_base.replace('_', ' ') in file_base or 
                                    group_base.replace(' ', '_') in file_base):
                                    excel_path = os.path.join(search_dir, filename)
                                    log.info("Found Excel file in same directory: %s", excel_path)
                                    return excel_path
                
                # ПРИОРИТЕТ 2: Ищем Excel файл во всех поддиректориях hosting/open
                root = os.path.dirname(os.path.abspath(__file__))
                hosting_open = os.path.join(root, "hosting", "open")
                if os.path.exists(hosting_open):
                    for ext in ['.xlsx', '.xls']:
                        patterns = [
                            os.path.join(hosting_open, "**", f"{group_base}{ext}"),
                            os.path.join(hosting_open, "**", f"{group_base.replace('_', ' ')}{ext}"),
                            os.path.join(hosting_open, "**", f"{group_base.replace(' ', '_')}{ext}"),
                        ]
                        for pattern in patterns:
                            matches = glob.glob(pattern, recursive=True)
                            if matches:
                                matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                                log.info("Found Excel file in hosting/open: %s", matches[0])
                                return matches[0]
                
                # ПРИОРИТЕТ 3: Ищем по оригинальному имени файла в uploads (последний вариант)
                uploads_dir = os.path.join(root, "uploads")
                if os.path.exists(uploads_dir):
                    for filename in os.listdir(uploads_dir):
                        if filename.lower().endswith(('.xlsx', '.xls')):
                            filename_base = os.path.splitext(filename)[0]
                            if (group_base in filename_base or filename_base in group_base or
                                group_base.replace('_', ' ') in filename_base or 
                                group_base.replace(' ', '_') in filename_base):
                                excel_path = os.path.join(uploads_dir, filename)
                                log.info("Found Excel file in uploads: %s", excel_path)
                                return excel_path
                
                return None

            def read_excel_cell(excel_path: str, row: int, col_letter: str):
                """Читает значение из ячейки Excel по буквенному обозначению столбца."""
                if not excel_path or not os.path.exists(excel_path):
                    log.warning("Excel file not found: %s", excel_path)
                    return None
                
                try:
                    # Пробуем использовать openpyxl
                    try:
                        import openpyxl
                        # Сначала пробуем с data_only=True (для вычисленных значений)
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        cell_address = f"{col_letter}{row}"
                        log.info("Attempting to read Excel cell %s from %s (active sheet: %s)", cell_address, excel_path, ws.title)
                        cell_value = ws[cell_address].value
                        log.info("Read Excel cell %s from %s (data_only=True): %s (type: %s)", cell_address, excel_path, cell_value, type(cell_value).__name__)
                        
                        # Если значение None, пробуем без data_only (для формул)
                        if cell_value is None:
                            log.info("Value is None with data_only=True, trying data_only=False")
                            wb.close()
                            wb = openpyxl.load_workbook(excel_path, data_only=False)
                            ws = wb.active
                            cell_value = ws[cell_address].value
                            log.info("Read Excel cell %s from %s (data_only=False): %s (type: %s)", cell_address, excel_path, cell_value, type(cell_value).__name__)
                            
                            # Если это формула, пробуем получить ее текст
                            if hasattr(ws[cell_address], 'data_type') and ws[cell_address].data_type == 'f':
                                formula = ws[cell_address].value
                                log.info("Cell %s contains formula: %s", cell_address, formula)
                        
                        # Если все еще None, пробуем другие листы
                        if cell_value is None and len(wb.sheetnames) > 1:
                            log.info("Value still None, checking other sheets: %s", wb.sheetnames)
                            for sheet_name in wb.sheetnames:
                                if sheet_name != ws.title:
                                    ws_test = wb[sheet_name]
                                    cell_value = ws_test[cell_address].value
                                    log.info("Checked sheet '%s', cell %s: %s", sheet_name, cell_address, cell_value)
                                    if cell_value is not None:
                                        log.info("Found value in sheet '%s': %s", sheet_name, cell_value)
                                        break
                        
                        # Проверяем соседние ячейки для отладки
                        if cell_value is None:
                            log.warning("Cell %s is None. Checking nearby cells for debugging:", cell_address)
                            for offset_row in [-1, 0, 1]:
                                for offset_col in [-1, 0, 1]:
                                    if offset_row == 0 and offset_col == 0:
                                        continue
                                    try:
                                        # Конвертируем смещение столбца
                                        col_num = ord(col_letter[-1]) - ord('A') + 1
                                        new_col_letter = chr(ord('A') + col_num + offset_col - 1)
                                        if col_num + offset_col > 0:
                                            test_address = f"{new_col_letter}{row + offset_row}"
                                            test_value = ws[test_address].value
                                            if test_value is not None:
                                                log.info("  Nearby cell %s: %s", test_address, test_value)
                                    except Exception:
                                        pass
                        
                        wb.close()
                        return cell_value
                    except ImportError:
                        log.warning("openpyxl not installed, using pandas fallback")
                        # Fallback на pandas
                        df = pd.read_excel(excel_path, header=None, engine='openpyxl' if 'openpyxl' in str(pd.read_excel.__defaults__) else None)
                        # Конвертируем букву столбца в индекс
                        # AS = A(1)*26 + S(19) = 26 + 19 = 45, но индексация с 0, так что 44
                        col_idx = 0
                        for char in col_letter:
                            col_idx = col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
                        col_idx -= 1  # Индексация с 0
                        log.debug("Converted column %s to index %d (row %d, df shape: %s)", col_letter, col_idx, row, df.shape)
                        if row - 1 < len(df) and col_idx < len(df.columns):
                            cell_value = df.iloc[row - 1, col_idx]
                            log.debug("Read Excel cell %s%d from %s via pandas: %s", col_letter, row, excel_path, cell_value)
                            return cell_value
                        else:
                            log.warning("Cell %s%d is out of bounds (df shape: %s, col_idx: %d)", col_letter, row, df.shape, col_idx)
                        return None
                    except Exception as e:
                        log.exception("Error reading Excel cell %s%d from %s with openpyxl: %s", col_letter, row, excel_path, e)
                        # Пробуем pandas как fallback
                        try:
                            df = pd.read_excel(excel_path, header=None)
                            col_idx = 0
                            for char in col_letter:
                                col_idx = col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
                            col_idx -= 1
                            if row - 1 < len(df) and col_idx < len(df.columns):
                                return df.iloc[row - 1, col_idx]
                        except Exception as e2:
                            log.exception("Error reading Excel cell %s%d from %s with pandas: %s", col_letter, row, excel_path, e2)
                        return None
                except Exception as e:
                    log.exception("Error reading Excel cell %s%d from %s: %s", col_letter, row, excel_path, e)
                    return None

            def parse_percent(value):
                """Парсит процентное значение из строки.
                Если значение < 1, считаем это десятичной дробью (0.8913 -> 89.13%).
                Если значение >= 1, считаем это процентами (96.67 -> 96.67% или 96.67% -> 96.67).
                """
                if value is None:
                    return None
                s = str(value).strip().replace('%', '').replace(',', '.')
                try:
                    val = float(s)
                    # Если значение меньше 1, это десятичная дробь (0.8913 = 89.13%)
                    if val < 1.0:
                        val = val * 100
                    return val
                except Exception:
                    return None
            
            def format_percent_display(value):
                """Форматирует процент для отображения (округление до 2 знаков)."""
                if value is None:
                    return "0%"
                try:
                    rounded = round(float(value), 2)
                    if rounded.is_integer():
                        return f"{int(rounded)}%"
                    return f"{rounded}%"
                except Exception:
                    return f"{value}%"

            # Используем personal_path если передан (тот же файл, что и для ведомости)
            # Иначе ищем через find_rr_csv
            csv_path = None
            if personal_path and os.path.exists(personal_path):
                csv_path = personal_path
                log.info("Using personal_path for RR calculation: %s", csv_path)
            elif file_name:
                csv_path = find_rr_csv(file_name, uid)
                if csv_path:
                    log.info("Found CSV via find_rr_csv: %s", csv_path)
            
            if not csv_path:
                log.error("CSV for RR not found: personal_path=%s, file_name=%s, uid=%s", personal_path, file_name, uid)
                raise FileNotFoundError("CSV for RR not found")
            
            try:
                df = pd.read_csv(csv_path, dtype=str)
            except Exception:
                df = pd.read_csv(csv_path, encoding='cp1251', dtype=str)
            if df is None or df.empty or 'vk_id' not in df.columns:
                raise ValueError("CSV missing data or vk_id column")

            def _norm_vk(val):
                s = str(val).strip()
                m = re.search(r'(\d+)', s)
                return m.group(1) if m else s

            df_valid = df[df['vk_id'].notna()].copy()
            df_valid['_vk_norm'] = df_valid['vk_id'].apply(_norm_vk)
            log.info("Looking for user %s in CSV %s. Found %d rows with valid vk_id", uid, csv_path, len(df_valid))
            log.info("Sample normalized vk_ids: %s", df_valid['_vk_norm'].head(10).tolist() if len(df_valid) > 0 else [])
            
            row = df_valid[df_valid['_vk_norm'].astype(str) == str(uid)]
            if row.empty:
                log.error("User %s not found in CSV %s. Available vk_ids (normalized): %s", uid, csv_path, df_valid['_vk_norm'].unique().tolist())
                raise ValueError("User row not found in CSV")
            
            log.info("Found user row in CSV. Row index: %s", row.index.tolist())
            p = row.fillna('0').iloc[0]
            
            # Логируем все значения из найденной строки
            log.info("Reading data from CSV row for user %s:", uid)
            log.info("  CSV file: %s", csv_path)
            log.info("  name: %s", p.get('name', ''))
            log.info("  vk_id: %s", p.get('vk_id', ''))
            log.info("  class: %s", p.get('class', ''))
            log.info("  type: %s", p.get('type', ''))
            log.info("  rr_gk: %s", p.get('rr_gk', ''))
            log.info("  rr_gkp: %s", p.get('rr_gkp', ''))
            log.info("  stud_gk: %s", p.get('stud_gk', ''))
            log.info("  stud_gkp: %s", p.get('stud_gkp', ''))
            log.info("  base: %s", p.get('base', ''))
            log.info("  rr_salary_gk: %s", p.get('rr_salary_gk', ''))
            log.info("  rr_salary_gkp: %s", p.get('rr_salary_gkp', ''))

            def to_float_safe(v):
                try:
                    s = str(v).replace(',', '.').replace('%', '').strip()
                    return float(s)
                except Exception:
                    return None

            def to_int_safe(v):
                try:
                    return int(float(str(v)))
                except Exception:
                    return 0

            # Читаем данные пользователя
            course_type = str(p.get('class', '')).strip()
            curator_type = str(p.get('type', '')).strip()
            rr_gk_str = p.get('rr_gk', '')
            rr_gkp_str = p.get('rr_gkp', '')
            stud_gk = to_int_safe(p.get('stud_gk', 0))
            stud_gkp = to_int_safe(p.get('stud_gkp', 0))
            base = to_int_safe(p.get('base', 0))
            rr_salary_gk = to_float_safe(p.get('rr_salary_gk', 0))
            rr_salary_gkp = to_float_safe(p.get('rr_salary_gkp', 0))
            
            log.info("RR calculation input: course_type=%s, curator_type=%s, rr_gk_str=%s, rr_gkp_str=%s, stud_gk=%s, stud_gkp=%s, base=%s, rr_salary_gk=%s, rr_salary_gkp=%s", 
                    course_type, curator_type, rr_gk_str, rr_gkp_str, stud_gk, stud_gkp, base, rr_salary_gk, rr_salary_gkp)

            # Находим Excel файл
            # Сначала находим директорию, где находится групповой CSV (hosting/open/тест/тест/7/)
            group_csv_dir = None
            if 'users' in csv_path:
                group_csv_dir = os.path.dirname(csv_path)  # hosting/open/тест/тест/7/users
                group_csv_dir = os.path.dirname(group_csv_dir)  # hosting/open/тест/тест/7
            else:
                group_csv_dir = os.path.dirname(csv_path)
            
            log.info("Looking for Excel file in directory: %s (CSV path: %s, file_name: %s)", group_csv_dir, csv_path, file_name)
            
            # Ищем Excel файл
            # ПРИОРИТЕТ 1: Ищем в директории группового CSV (hosting/open/тест/тест/7/)
            excel_path = None
            if group_csv_dir and os.path.exists(group_csv_dir):
                # Сначала пробуем найти по имени file_name (base_name) - это будет тест_тест_7.xlsx
                if file_name:
                    for ext in ['.xlsx', '.xls']:
                        test_path = os.path.join(group_csv_dir, file_name + ext)
                        if os.path.exists(test_path):
                            excel_path = test_path
                            log.info("Found Excel file by file_name in group CSV directory: %s", excel_path)
                            break
                
                # Если не нашли по file_name, ищем любой Excel файл в этой директории
                if not excel_path:
                    excel_files = []
                    if os.path.exists(group_csv_dir):
                        for filename in os.listdir(group_csv_dir):
                            if filename.lower().endswith(('.xlsx', '.xls')) and 'users' not in filename:
                                excel_files.append(os.path.join(group_csv_dir, filename))
                    
                    if excel_files:
                        # Берем самый свежий файл
                        excel_files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                        excel_path = excel_files[0]
                        log.info("Found Excel file (any) in group CSV directory: %s", excel_path)
            
            # ПРИОРИТЕТ 2: Если не нашли, ищем в uploads по исходному имени (может быть тест.xlsx)
            if not excel_path:
                root = os.path.dirname(os.path.abspath(__file__))
                uploads_dir = os.path.join(root, "uploads")
                if os.path.exists(uploads_dir):
                    # Ищем любой Excel файл в uploads (возможно, исходный файл)
                    uploads_excel = []
                    for filename in os.listdir(uploads_dir):
                        if filename.lower().endswith(('.xlsx', '.xls')):
                            uploads_excel.append(os.path.join(uploads_dir, filename))
                    
                    if uploads_excel:
                        # Берем самый свежий файл из uploads
                        uploads_excel.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                        excel_path = uploads_excel[0]
                        log.info("Found Excel file in uploads: %s", excel_path)
            
            # ПРИОРИТЕТ 3: Используем функцию поиска
            if not excel_path:
                log.info("Excel file not found in %s or uploads, using find_excel_file function", group_csv_dir)
                excel_path = find_excel_file(csv_path)
            
            if not excel_path:
                log.warning("Excel file not found for CSV: %s. Searched in: %s, file_name: %s", csv_path, group_csv_dir, file_name)
                raise FileNotFoundError("Excel file not found")
            
            log.info("Using Excel file for RR calculation: %s (CSV: %s)", excel_path, csv_path)

            reply_parts = []
            reply_parts.append("Твой RR считается по формуле из договора, а именно:")

            # Определяем, для каких курсов нужно показать расчет
            show_gk = course_type == 'ГК' and stud_gk > 0
            show_gkp = course_type == 'ГК+' and stud_gkp > 0
            
            # Для ГК/ГК+ показываем оба
            if course_type == 'ГК/ГК+':
                show_gk = stud_gk > 0
                show_gkp = stud_gkp > 0
            
            log.info("RR calculation: course_type=%s, curator_type=%s, stud_gk=%d, stud_gkp=%d, show_gk=%s, show_gkp=%s", 
                    course_type, curator_type, stud_gk, stud_gkp, show_gk, show_gkp)

            # Определяем строки для чтения min/max (только строки 3 и 11)
            row_min_max = 3 if curator_type == 'Личный' else 11  # Для AS3/AT3 или AS11/AT11
            log.info("Reading min/max from Excel row %d for curator_type=%s", row_min_max, curator_type)

            # Флаг для отслеживания, была ли добавлена формула расчета
            has_formula = False

            for course_name, is_gk in [('ГК', True), ('ГК+', False)]:
                # Пропускаем, если для этого курса не нужно показывать расчет
                if (is_gk and not show_gk) or (not is_gk and not show_gkp):
                    log.debug("Skipping %s: is_gk=%s, show_gk=%s, show_gkp=%s", course_name, is_gk, show_gk, show_gkp)
                    continue
                
                log.info("Processing RR calculation for %s", course_name)

                # Определяем столбцы для min/max
                if is_gk:
                    min_col = 'AS'
                    max_col = 'AT'
                    min_contract_col = 'AS'  # Для проверки минимального значения по договору
                    rr_value_str = rr_gk_str
                    stud_count = stud_gk
                    rr_salary = rr_salary_gk
                else:
                    min_col = 'AW'
                    max_col = 'AX'
                    min_contract_col = 'AW'  # Для проверки минимального значения по договору
                    rr_value_str = rr_gkp_str
                    stud_count = stud_gkp
                    rr_salary = rr_salary_gkp

                log.info("Processing %s: rr_value_str=%s, stud_count=%d, base=%s", course_name, rr_value_str, stud_count, base)
                
                if not rr_value_str or stud_count == 0:
                    log.warning("Skipping %s: rr_value_str=%s, stud_count=%d", course_name, rr_value_str, stud_count)
                    continue
                
                if base == 0:
                    log.warning("Skipping %s: base is 0", course_name)
                    continue

                # Читаем min/max из Excel (строки 3 или 11)
                log.info("Reading min/max for %s: %s%d and %s%d", course_name, min_col, row_min_max, max_col, row_min_max)
                min_val = read_excel_cell(excel_path, row_min_max, min_col)
                max_val = read_excel_cell(excel_path, row_min_max, max_col)
                log.info("Read values for %s: min=%s (from %s%d), max=%s (from %s%d)", 
                        course_name, min_val, min_col, row_min_max, max_val, max_col, row_min_max)

                if min_val is None or max_val is None:
                    log.warning("Could not read min/max from Excel for %s: min=%s (from %s%d), max=%s (from %s%d). Excel file: %s", 
                              course_name, min_val, min_col, row_min_max, max_val, max_col, row_min_max, excel_path)
                    continue

                try:
                    min_raw = float(str(min_val).replace(',', '.').replace('%', ''))
                    max_raw = float(str(max_val).replace(',', '.').replace('%', ''))
                    # Если значения меньше 1, считаем это десятичной дробью и преобразуем в проценты
                    min_float = min_raw * 100 if min_raw < 1.0 else min_raw
                    max_float = max_raw * 100 if max_raw < 1.0 else max_raw
                    log.info("Parsed min/max for %s: min_raw=%s -> min_float=%s, max_raw=%s -> max_float=%s", 
                            course_name, min_raw, min_float, max_raw, max_float)
                except Exception as e:
                    log.warning("Could not parse min/max values: min=%s, max=%s, error=%s", min_val, max_val, e)
                    continue

                # Парсим RR значение (процент)
                log.info("Parsing RR value for %s: rr_value_str=%s", course_name, rr_value_str)
                rr_percent = parse_percent(rr_value_str)
                if rr_percent is None:
                    log.warning("Could not parse RR value: %s", rr_value_str)
                    continue
                log.info("Parsed RR percent for %s: %s", course_name, rr_percent)
                
                # Проверяем, если RR <= минимального значения по договору
                if rr_percent <= min_float:
                    log.info("RR %s <= min %s for %s, showing special message", rr_percent, min_float, course_name)
                    rr_display = format_percent_display(rr_percent)
                    min_display = format_percent_display(min_float)
                    reply_parts.append(f"\n\nДля {course_name}: Твой RR, взятый из ЖО: {rr_display} меньше или равен минимальному значению по договору: {min_display}, поэтому оплата за RR составляет: {rr_salary:.2f}₽")
                    continue

                # Если значение ФП больше максимума, то берется максимум
                fp_value = rr_percent
                fp_exceeds_max = False
                if rr_percent > max_float:
                    log.info("RR %s > max %s for %s, using max value", rr_percent, max_float, course_name)
                    fp_value = max_float
                    fp_exceeds_max = True

                # Определяем, какой тип куратора (сотник или личный)
                is_sotnik = curator_type != 'Личный'
                
                # Рассчитываем по формуле в зависимости от типа куратора
                if is_sotnik:
                    # Для сотников: (ФП-МИН)/(МАКС-МИН)*0.7*КОЛ-ВО_ДЕТЕЙ*30
                    log.info("Calculating RR for %s (сотник): fp_value=%s (rr_percent=%s), min_float=%s, max_float=%s, stud_count=%s", 
                            course_name, fp_value, rr_percent, min_float, max_float, stud_count)
                    
                    if max_float - min_float == 0:
                        calculated = 0
                        log.warning("max_float - min_float == 0 for %s, setting calculated=0", course_name)
                    else:
                        calculated = ((fp_value - min_float) / (max_float - min_float)) * 0.7 * stud_count * 30
                        log.info("Calculated RR for %s (сотник): %s", course_name, calculated)

                    # Округляем значения для отображения в формуле
                    fp_display_val = round(fp_value, 2)
                    min_display_val = round(min_float, 2)
                    max_display_val = round(max_float, 2)
                    
                    formula_text = f"\n\nДля {course_name}: Оплата за RR = (ФП-МИН) / (МАКС-МИН)* 0.7 * КОЛ-ВО_ДЕТЕЙ * 30 =({fp_display_val} - {min_display_val})/({max_display_val} - {min_display_val})*0.7*{stud_count}*30 = {calculated:.2f}₽"
                    
                    # Добавляем примечание, если ФП превысил максимум
                    if fp_exceeds_max:
                        formula_text += "\n*в данном случае ФП=МАКС, так как твой показатель ФП превысил максимальный процент (подробнее см. в договоре)"
                else:
                    # Для личных: (ФП-МИН)/(МАКС-МИН)*0.3*0.7*КОЛ-ВО_ДЕТЕЙ*СТАВКА_ЗА_УЧЕНИКА
                    log.info("Calculating RR for %s (личный): fp_value=%s (rr_percent=%s), min_float=%s, max_float=%s, stud_count=%s, base=%s", 
                            course_name, fp_value, rr_percent, min_float, max_float, stud_count, base)
                    
                    if max_float - min_float == 0:
                        calculated = 0
                        log.warning("max_float - min_float == 0 for %s, setting calculated=0", course_name)
                    else:
                        calculated = ((fp_value - min_float) / (max_float - min_float)) * 0.3 * 0.7 * stud_count * base
                        log.info("Calculated RR for %s (личный): %s", course_name, calculated)

                    # Округляем значения для отображения в формуле
                    fp_display_val = round(fp_value, 2)
                    min_display_val = round(min_float, 2)
                    max_display_val = round(max_float, 2)
                    
                    formula_text = f"\n\nДля {course_name}: Оплата за RR = (ФП-МИН) / (МАКС-МИН)*0.3 * 0.7 * КОЛ-ВО_ДЕТЕЙ * СТАВКА_ЗА_УЧЕНИКА =({fp_display_val} - {min_display_val})/({max_display_val} - {min_display_val})*0.3*0.7*{stud_count}*{base} = {calculated:.2f}₽"
                    
                    # Добавляем примечание, если ФП превысил максимум
                    if fp_exceeds_max:
                        formula_text += "\n*в данном случае ФП=МАКС, так как твой показатель ФП превысил максимальный процент (подробнее см. в договоре)"
                
                reply_parts.append(formula_text)
                has_formula = True

            log.info("Finished RR calculation loop. reply_parts length: %d, has_formula=%s", len(reply_parts), has_formula)
            if len(reply_parts) == 1:
                # Не удалось рассчитать ни для одного курса (только начальное сообщение)
                log.error("Could not calculate RR for any course. reply_parts: %s", reply_parts)
                raise ValueError("Could not calculate RR for any course")

            # Добавляем пояснение только если была добавлена формула расчета
            if has_formula:
                is_sotnik = curator_type != 'Личный'
                reply_parts.append("\n\nФП - фактическое значение RR, взятое из ЖО за ПРОШЛЫЙ блок.")
                reply_parts.append("\nМАКС, МИН - Макс. и мин. проценты из договора.")
                if not is_sotnik:
                    reply_parts.append("\n0.3 - Добавочный коэффициент.")
                reply_parts.append("\n0.7 - Вес метрики.")
                reply_parts.append("\nКОЛ-ВО_ДЕТЕЙ - Количество детей в твоей группе.")
                if is_sotnik:
                    reply_parts.append("\n30 - Фиксированная оплата за одного обучающегося.")
                else:
                    reply_parts.append("\nСТАВКА_ЗА_УЧЕНИКА - фиксированная оплата за одного обучающегося.")

            reply = "".join(reply_parts)

        except Exception as e:
            log.exception("Error processing RR calculation: %s", e)
            reply = (f"Данные по retention rate взяты из журнала оплат за предыдущий блок (например, если мы считаем выплату за 2-й блок, то берём RR с 1 на 2 блок."
                     f"\nВсе причины слива одинаково учитываются в Retention Rate. Если произошло обстоятельство непреодолимой силы (например, ученик погиб), ты можешь обратиться к СК для корректировки RR, но только в таких случаях")
    else:
            reply = ""

    return reply

def _find_curator_csv(base_name: str, vk_uid: int):
    root = os.path.dirname(os.path.abspath(__file__))
    patterns = [
        os.path.join(root, "hosting", "open", "**", "users", f"{vk_uid}_*_{base_name}.csv"),
        os.path.join(root, "hosting", "open", "**", "users", f"{vk_uid}_*_{base_name.replace(' ','_')}.csv"),
    ]
    matches = []
    for pat in patterns:
        matches.extend(glob.glob(pat, recursive=True))
    if matches:
        matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return matches[0]
    group_patterns = [
        os.path.join(root, "hosting", "open", "**", f"{base_name}.csv"),
        os.path.join(root, "hosting", "open", "**", f"{base_name.replace(' ','_')}.csv"),
    ]
    group_matches = []
    for pat in group_patterns:
        group_matches.extend(glob.glob(pat, recursive=True))
    if group_matches:
        group_matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return group_matches[0]
    return None


def _to_int_safe(value) -> int:
    try:
        if value is None:
            return 0
        s = str(value).replace('\u00A0', '').replace('\xa0', '').replace(' ', '').replace(',', '.')
        if s == '' or s.lower() == 'nan':
            return 0
        return int(float(s))
    except Exception:
        return 0

def _to_float_safe(value) -> float:
    try:
        if value is None:
            return 0.0
        s = str(value).replace('\u00A0', '').replace('\xa0', '').replace(' ', '').replace(',', '.')
        if s == '' or s.lower() == 'nan':
            return 0.0
        return float(s)
    except Exception:
        return 0.0

def _to_float_str_money(value) -> str:
    try:
        if value is None:
            return '0'
        s = str(value).replace('\u00A0', '').replace('\xa0', '').replace(' ', '').replace(',', '.')
        if s == '' or s.lower() == 'nan':
            return '0'
        return str(round(float(s), 2))
    except Exception:
        return '0'

def _format_payment_label(original_filename: str, idx: int, max_length: int = 30, created_at: float = None, db_id: int = None, groups: str = None) -> str:
    """Форматирует название выплаты для кнопки, убирая расширение .csv и ограничивая длину"""
    if original_filename:
        # Убираем расширение .csv
        base_name = os.path.splitext(original_filename)[0]
        
        # Добавляем информацию о группах для различения
        if groups and groups.strip():
            # Извлекаем краткое название группы (например "1" из "Аня Колотович | Группа 1")
            group_info = groups.strip()
            if '|' in group_info:
                group_part = group_info.split('|')[-1].strip()
                if group_part:
                    # Убираем слово "Группа" и берем только номер/название
                    if group_part.lower().startswith('группа '):
                        group_clean = group_part[7:]  # Убираем "Группа "
                        if group_clean:
                            base_name = f"{base_name} ({group_clean})"
                    else:
                        base_name = f"{base_name} ({group_part})"
            else:
                # Если нет разделителя, берем последние 10 символов
                group_short = group_info[-10:] if len(group_info) > 10 else group_info
                base_name = f"{base_name} ({group_short})"
        
        # Если есть временная метка, добавляем её для дополнительного различия
        elif created_at and db_id:
            import time
            try:
                # Форматируем дату как день/месяц
                date_str = time.strftime('%d.%m', time.localtime(created_at))
                base_name = f"{base_name} ({date_str})"
            except Exception:
                # Fallback - используем db_id
                base_name = f"{base_name} #{db_id}"
        
        # Ограничиваем длину
        if len(base_name) > max_length:
            base_name = base_name[:max_length-3] + "..."
            
        return base_name
    else:
        return f"Ведомость {idx}"

def _extract_numeric_vk(value: str) -> str:
    """Возвращает числовой VK ID из строки ссылки/ID."""
    if value is None:
        return ''
    s = str(value).strip()
    if not s:
        return ''
    match = re.search(r'(?:vk\.com/id|id)(\d{5,})', s, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r'(\d{5,})', s)
    if match:
        return match.group(1)
    return s

def _is_meaningful_comment(comment) -> bool:
    """Проверяет, содержит ли комментарий полезный текст (не '0', '-', 'nan' и т.п.)."""
    if comment is None:
        return False
    comment_str = str(comment).strip()
    if not comment_str:
        return False
    lowered = comment_str.lower()
    meaningless_values = {
        '0', '0.0', '0,0', 'nan', 'none', 'нет', 'no', 'none', 'пусто', 'n/a', 'н/д', '—', '-', '––'
    }
    if lowered in meaningless_values:
        return False
    if re.fullmatch(r'0+(\.0+)?', lowered):
        return False
    stripped = comment_str.strip('-').strip('—').strip()
    if not stripped:
        return False
    return True

def _compose_payment_sections(p: dict) -> dict:
    """Строит текстовые блоки для разных разделов выплаты."""
    sections = {
        "studs": "",
        "retention": "",
        "okk": "",
        "checks": "",
        "extras": "",
        "fines": "",
        "total": ""
    }

    def _as_money(value) -> str:
        return _to_float_str_money(value)

    def _format_percent(value) -> str:
        if value is None:
            return ''
        s = str(value).strip()
        if not s or s.lower() in ('nan', 'none', '-', '—'):
            return ''
        try:
            has_pct = s.endswith('%')
            num = float(s.replace('%', '').replace(',', '.'))
            if not has_pct and abs(num) <= 1:
                num *= 100
            num = round(num, 2)
            if num.is_integer():
                return f"{int(num)}%"
            return f"{num}%"
        except Exception:
                return s

    # Сопровождение ГК / ГК+
    stud_gk = _to_int_safe(p.get('stud_gk'))
    stud_gkp = _to_int_safe(p.get('stud_gkp'))
    stud_all = _to_int_safe(p.get('stud_all'))
    stud_rep = _to_int_safe(p.get('stud_rep'))
    rep_salary = _to_float_str_money(p.get('rep_salary'))
    base_val = _to_int_safe(p.get('base'))
    stud_salary_total = _to_int_safe(p.get('stud_salary'))
    raw_stud_salary_gk = p.get('stud_salary_gk')
    raw_stud_salary_gkp = p.get('stud_salary_gkp')
    stud_salary_gk_val = _to_int_safe(raw_stud_salary_gk)
    stud_salary_gkp_val = _to_int_safe(raw_stud_salary_gkp)

    slivs_gk = _to_int_safe(p.get('slivs_gk'))
    rr_gk = _format_percent(p.get('rr_gk'))
    rr_salary_gk = _as_money(p.get('rr_salary_gk'))
    okk_gk = _format_percent(p.get('okk_gk'))
    okk_salary_gk = _as_money(p.get('okk_salary_gk'))

    slivs_gkp = _to_int_safe(p.get('slivs_gkp'))
    rr_gkp = _format_percent(p.get('rr_gkp'))
    rr_salary_gkp = _as_money(p.get('rr_salary_gkp'))
    okk_gkp = _format_percent(p.get('okk_gkp'))
    okk_salary_gkp = _as_money(p.get('okk_salary_gkp'))

    kpi_total_val = _to_float_str_money(p.get('kpi_total'))
    splitted_blocks = []

    def _append_block(title: str, lines) -> None:
        if lines:
            splitted_blocks.append(f"\n{title}\n" + ''.join(lines) + "\n")

    def _has_explicit(value) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() != ''
        return True

    gk_lines = []
    if stud_gk > 0:
        gk_lines.append(f"\nВсего учеников - ГК: {stud_gk}")
    if base_val > 0:
        gk_lines.append(f"\nОклад за ученика: {base_val}₽")
    if stud_salary_gk_val > 0 or _has_explicit(raw_stud_salary_gk):
        gk_lines.append(f"\n→ Сумма оклада: {stud_salary_gk_val}₽")
    elif stud_salary_total > 0:
        gk_lines.append(f"\n→ Сумма оклада: {stud_salary_total}₽")
        if stud_rep > 0:
            gk_lines.append(f"\nКол-во учеников с тарифом с репетитором: {stud_rep}")
    if rep_salary != '0' and _to_float_safe(rep_salary) > 0:
        gk_lines.append(f"\nДоплата за учеников с репетитором: {rep_salary}₽")
    if rr_gk:
        gk_lines.append(f"\nRetention ГК: {rr_gk}")
    if rr_salary_gk != '0':
        gk_lines.append(f"\n→ Оплата за retention ГК: {rr_salary_gk}₽")
    if okk_gk:
        gk_lines.append(f"\nOKK ГК: {okk_gk}")
    if okk_salary_gk != '0':
        gk_lines.append(f"\n→ Оплата за OKK ГК: {okk_salary_gk}₽")
    
    # Calculate KPI for GK - выводим всегда, если есть хотя бы одна из строк оплаты
    if rr_salary_gk != '0' or okk_salary_gk != '0':
        kpi_gk = _to_float_safe(rr_salary_gk) + _to_float_safe(okk_salary_gk)
        kpi_gk_str = _to_float_str_money(kpi_gk)
        gk_lines.append(f"\n→ Сумма KPI (OKK+Retention): {kpi_gk_str}₽")
    
    # Append block only if there are students (stud_gk > 0)
    if stud_gk > 0:
        _append_block('[Сопровождение ГК]', gk_lines)

    gkp_lines = []
    if stud_gkp > 0:
        gkp_lines.append(f"\nВсего учеников - ГК+: {stud_gkp}")
    if base_val > 0:
        gkp_lines.append(f"\nОклад за ученика: {base_val}₽")
    if stud_salary_gkp_val > 0 or _has_explicit(raw_stud_salary_gkp):
        gkp_lines.append(f"\n→ Сумма оклада: {stud_salary_gkp_val}₽")
    if rr_gkp:
        gkp_lines.append(f"\nRetention ГК+: {rr_gkp}")
    if rr_salary_gkp != '0':
        gkp_lines.append(f"\n→ Оплата за retention ГК+: {rr_salary_gkp}₽")
    if okk_gkp:
        gkp_lines.append(f"\nOKK ГК+: {okk_gkp}")
    if okk_salary_gkp != '0':
        gkp_lines.append(f"\n→ Оплата за OKK ГК+: {okk_salary_gkp}₽")
    
    # Calculate KPI for GK+ - выводим всегда, если есть хотя бы одна из строк оплаты
    if rr_salary_gkp != '0' or okk_salary_gkp != '0':
        kpi_gkp = _to_float_safe(rr_salary_gkp) + _to_float_safe(okk_salary_gkp)
        kpi_gkp_str = _to_float_str_money(kpi_gkp)
        gkp_lines.append(f"\n→ Сумма KPI (OKK+Retention): {kpi_gkp_str}₽")
    
    # Append block only if there are students (stud_gkp > 0)
    if stud_gkp > 0:
        _append_block('[Сопровождение ГК+]', gkp_lines)

    if splitted_blocks:
        studs_combined = ''.join(splitted_blocks)
        sections['studs'] = studs_combined
    elif any(value > 0 for value in (stud_all, stud_rep, stud_salary_total)) or (rep_salary != '0' and _to_float_safe(rep_salary) > 0):
        # Проверяем, есть ли хотя бы одно ненулевое значение (кроме base_val и kpi_total_val)
        has_meaningful_data = any(value > 0 for value in (stud_all, stud_rep, stud_salary_total)) or (rep_salary != '0' and _to_float_safe(rep_salary) > 0)
        if has_meaningful_data:
            studs_section = "\n[Сопровождение учеников]"
            if stud_all > 0:
                studs_section += f"\nВсего учеников в группах: {stud_all}"
            if stud_rep > 0:
                studs_section += f"\nКол-во учеников с тарифом с репетитором: {stud_rep}"
            if base_val > 0:
                studs_section += f"\nОклад за ученика: {base_val}₽"
            if rep_salary != '0' and _to_float_safe(rep_salary) > 0:
                studs_section += f"\nДоплата за учеников с репетитором: {rep_salary}₽"
            if stud_salary_total > 0:
                studs_section += f"\n→ Сумма оклада: {stud_salary_total}₽"
            if kpi_total_val != '0' and _to_float_safe(kpi_total_val) > 0:
                studs_section += f"\n→ Сумма KPI (OKK+Retention): {kpi_total_val}₽"
            sections['studs'] = studs_section

    # Retention и OKK (общие блоки показываем только если нет отдельных секций)
    if not splitted_blocks:
        # Проверяем, есть ли хотя бы одно ненулевое значение в retention
        has_retention_data = (
            (rr_salary_gk != '0' and _to_float_safe(rr_salary_gk) > 0) or
            (rr_salary_gkp != '0' and _to_float_safe(rr_salary_gkp) > 0)
        )
        if has_retention_data:
            retention_section = "\n[Retention]"
            if rr_gk:
                retention_section += f"\nRetention ГК: {rr_gk}"
            if rr_salary_gk != '0':
                retention_section += f"\nОплата за retention ГК: {rr_salary_gk}₽"
            if rr_gkp:
                retention_section += f"\nRetention ГК+: {rr_gkp}"
            if rr_salary_gkp != '0':
                retention_section += f"\nОплата за retention ГК+: {rr_salary_gkp}₽"
            sections['retention'] = retention_section

        # Проверяем, есть ли хотя бы одно ненулевое значение в OKK
        has_okk_data = (
            (okk_salary_gk != '0' and _to_float_safe(okk_salary_gk) > 0) or
            (okk_salary_gkp != '0' and _to_float_safe(okk_salary_gkp) > 0) or
            (kpi_total_val != '0' and _to_float_safe(kpi_total_val) > 0)
        )
        if has_okk_data:
            okk_section = "\n[Показатели ОКК]"
            if okk_gk:
                okk_section += f"\nOKK ГК: {okk_gk}"
            if okk_salary_gk != '0':
                okk_section += f"\nОплата за OKK ГК: {okk_salary_gk}₽"
            if okk_gkp:
                okk_section += f"\nOKK ГК+: {okk_gkp}"
            if okk_salary_gkp != '0':
                okk_section += f"\nОплата за OKK ГК+: {okk_salary_gkp}₽"
            if kpi_total_val != '0' and _to_float_safe(kpi_total_val) > 0:
                okk_section += f"\n→ Сумма KPI (OKK+Retention): {kpi_total_val}₽"
            sections['okk'] = okk_section

    # Проверки (существующий блок)
    checks_section = ""
    checks_salary = _to_int_safe(p.get('checks_salary'))
    dop_checks = _to_int_safe(p.get('dop_checks'))
    if checks_salary > 0 or dop_checks > 0:
        if checks_salary > 0:
            checks_section += f"\n→ Проверка домашних работ: {checks_salary}₽"
        if dop_checks > 0:
            checks_section += f"\n→ Дополнительно – за проверки (данные СК): {dop_checks}₽"
        checks_section += "\n"
    sections["checks"] = checks_section

    # Дополнительная деятельность (существующий блок)
    extras_keys = ['up','chats','webs','meth','dop_sk','callsg','callsp']
    extras_names = {
        'up': 'За учебную поддержку',
        'chats': 'Модерация чатов',
        'webs': 'Модерация вебинаров',
        'callsg': 'Групповые созвоны',
        'callsp': 'Индивидуальные созвоны',
        'dop_sk': 'Доп. суммы, начисленные СК',
        'meth': 'Стол заказов',
    }
    extras_total = sum(_to_int_safe(p.get(k)) for k in extras_keys)
    dops_section = ""
    if extras_total > 0:
        dops_section = "\n\n[Иная деятельность]"
        for k in extras_keys:
            v = _to_int_safe(p.get(k))
            if v > 0:
                dops_section += f"\n{extras_names[k]}: {v}₽"
        dops_section += f"\n→ Всего в категории: {extras_total}₽"
    sections["extras"] = dops_section

    # Штрафы (существующий блок)
    fines_val = _to_int_safe(p.get('fines'))
    fines_section = f"\n\n→ Штрафы: -{fines_val}₽" if fines_val > 0 else f"\n\nШтрафы: отсутствуют"
    sections["fines"] = fines_section

    # Итого + комментарий
    total_section = f"\n\n→ ИТОГО К ВЫПЛАТЕ: {_to_float_str_money(p.get('total'))}₽"
    comment = p.get('comment', '')
    if _is_meaningful_comment(comment):
        total_section += f"\n[!!!] Комментарий: {str(comment).strip()}"
    sections["total"] = total_section

    return sections

def refresh_payment_status_from_db(payment_entry, user_id: int = None):
    """Обновляет статус выплаты из базы данных и возвращает актуальное значение."""
    if not os.path.exists(DB_PATH):
        return payment_entry.get("status")
    payment_id = payment_entry.get("id") or ""
    original_payment_id = payment_entry.get("original_payment_id") or ""
    db_id = payment_entry.get("db_id")
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        row = None
        if db_id:
            c.execute("SELECT status FROM vedomosti_users WHERE id = ?", (int(db_id),))
            row = c.fetchone()
        else:
            state_value = None
            if original_payment_id:
                state_value = f"imported:{original_payment_id}"
            elif '_' in payment_id:
                state_value = f"imported:{payment_id.split('_')[0]}"
            elif payment_id:
                state_value = f"imported:{payment_id}"
            if state_value:
                if user_id is not None:
                    c.execute("SELECT status, id FROM vedomosti_users WHERE vk_id = ? AND state = ?", (str(user_id), state_value))
                else:
                    c.execute("SELECT status, id FROM vedomosti_users WHERE state = ?", (state_value,))
                row = c.fetchone()
        conn.close()
        if row:
            if len(row) == 2 and not db_id:
                payment_entry["db_id"] = row[1]
                payment_entry["status"] = row[0] or ''
            else:
                payment_entry["status"] = row[0] or ''
        return payment_entry.get("status")
    except Exception:
        log.exception("Failed to refresh payment status from DB for payment_id=%s", payment_id)
        return payment_entry.get("status")

def format_message(file_name, uid, course_type, deadline):
    csv_path = _find_curator_csv(file_name, uid)
    if not csv_path:
        raise FileNotFoundError("CSV for curator not found")
    try:
        df = pd.read_csv(csv_path, dtype=str)
    except Exception:
        df = pd.read_csv(csv_path, encoding='cp1251', dtype=str)
    if df is None or df.empty:
        raise ValueError("CSV is empty")
    if 'vk_id' not in df.columns:
        raise ValueError("CSV missing vk_id column")
    uid_str = str(uid).strip()
    vk_series = df['vk_id'].fillna('').astype(str).apply(_extract_numeric_vk)
    row = df[vk_series == uid_str]
    if row.empty:
        raise ValueError("Curator vk_id not found in CSV")
    p = row.fillna('0').iloc[0]

    base = (f"=== Согласование выплаты ==="
            f"\nКурс: {course_type}"
            f"\nКуратор: {p.get('name','')}"
            f"\nТип куратора: {p.get('type','')}"
            f"\nПочта на платформе: {p.get('email','')}\n")

    sections = _compose_payment_sections(p)
    
    final = ("\n\nНажмите «Согласен», если у Вас нет разногласий с выставленными цифрами"
             "\nНажмите «Не согласен», если Вы не согласны с каким-либо из пунктов"
             f"\nДедлайн по согласованию выплаты: {deadline}")

    msg = (base
           + sections["studs"]
           + sections["retention"]
           + sections["okk"]
           + sections["checks"]
           + sections["extras"]
           + sections["fines"]
           + sections["total"]
           + final)
    phone = p.get('phone', '')
    console = p.get('console', '')
    return (msg, phone, console)

def add_payment_for_user(user_id: int, payment_data: dict) -> str:
    """Добавляет выплату в память и возвращает payment_id"""
    pid = str(uuid.uuid4())
    entry = {"id": pid, "data": payment_data, "created_at": time.time(), "status": "new"}
    with user_payments_lock:
        user_payments.setdefault(user_id, []).append(entry)
    log.info("Добавлена выплата %s для user %s (fio=%s file=%s)", pid, user_id, payment_data.get('fio',''), payment_data.get('original_filename',''))
    return pid


def get_payments_for_user(user_id: int):
    with user_payments_lock:
        return user_payments.get(user_id, []).copy()  # Возвращаем копию для безопасности


def get_all_payments_for_user_from_db(user_id: int, limit: int = 100):
    """Загружает ВСЕ ведомости пользователя из базы данных, не только те что в памяти.
    Возвращает список в том же формате что и get_payments_for_user."""
    if not os.path.exists(DB_PATH):
        log.warning("DB file not found: %s", DB_PATH)
        return []
    
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        # Получаем все ведомости пользователя, отсортированные по времени создания (новые сначала)
        c.execute("""
            SELECT id, vk_id, personal_path, original_filename, state, status, disagree_reason, confirmed_at, created_at 
            FROM vedomosti_users 
            WHERE vk_id = ?
              AND (state LIKE 'imported:%' OR state LIKE 'repet_imported:%')
            ORDER BY created_at DESC, id DESC
            LIMIT ?
        """, (str(user_id), limit))
        rows = c.fetchall()
        conn.close()
        
        payments = []
        for db_row in rows:
            try:
                db_id, vk_id_raw, personal_path, original_filename, state, status_db, disagree_reason_db, confirmed_at_db, created_at_db = db_row
                
                if not state or (not state.startswith('imported:') and not state.startswith('repet_imported:')):
                    continue
                
                # Определяем тип выплаты
                is_repet = state.startswith('repet_imported:')
                prefix = 'repet_imported:' if is_repet else 'imported:'
                
                parts = state.split(prefix, 1)
                if len(parts) != 2 or not parts[1]:
                    continue
                
                payment_id = parts[1]
                
                # Создаем уникальный payment_id на основе db_id для старых записей с дублирующимися payment_id
                unique_payment_id = f"{payment_id}_{db_id}" if payment_id else f"payment_{db_id}"
                
                # ВСЕГДА загружаем данные из CSV файла для каждой записи БД
                # Не используем кэш памяти, так как у одного пользователя могут быть разные ведомости
                row_dict = {}
                if personal_path and os.path.exists(personal_path):
                    try:
                        df = get_cached_csv_data(personal_path)
                        if isinstance(df, pd.DataFrame) and not df.empty:
                            row_dict = df.iloc[0].to_dict()
                        log.debug("Loaded CSV data for payment %s from %s", unique_payment_id, personal_path)
                    except Exception:
                        log.warning("Failed to read CSV for payment %s path=%s", unique_payment_id, personal_path)
                else:
                    log.warning("Personal path not found for payment %s: %s", unique_payment_id, personal_path)
                
                # Для репетиторов используем отдельный маппер
                if is_repet:
                    payment_data = _map_row_to_repet_payment_data(row_dict, user_id, original_filename)
                    payment_data['is_repet'] = True
                else:
                    payment_data = _map_row_to_payment_data(row_dict, user_id, original_filename)
                    payment_data['is_repet'] = False
                
                # ВАЖНО: Добавляем personal_path в данные, чтобы форматтер знал источник
                payment_data['personal_path'] = personal_path
                
                entry = {
                    "id": unique_payment_id,
                    "data": payment_data,
                    "created_at": float(created_at_db) if created_at_db else time.time(),
                    "status": status_db or "new",
                    "db_id": db_id,  # Сохраняем db_id для отладки
                    "original_payment_id": payment_id  # Сохраняем исходный payment_id
                }
                
                if disagree_reason_db:
                    entry["disagree_reason"] = disagree_reason_db
                    
                payments.append(entry)
                    
            except Exception:
                log.exception("Error loading payment from DB row %s", db_row)
        
        log.info("Loaded %d total payments for user %s from DB (limit=%d)", len(payments), user_id, limit)
        return payments
        
    except Exception:
        log.exception("Failed to load all payments for user %s from DB", user_id)
        return []


def find_payment(user_id: int, payment_id: str):
    log.debug("find_payment called: user_id=%s, payment_id=%s", user_id, payment_id)
    # Сначала ищем в памяти (быстрее и актуальнее)
    with user_payments_lock:
        for p in user_payments.get(user_id, []):
            if p["id"] == payment_id:
                log.debug("Found payment %s in memory for user %s", payment_id, user_id)
                return p
    
    # Если не нашли в памяти, ищем в базе данных
    try:
        if not os.path.exists(DB_PATH):
            return None
        
        conn = sqlite3.connect(DB_PATH, timeout=30)
        c = conn.cursor()
        
        # Сначала пробуем найти по полному уникальному payment_id (новый формат)
        if '_' in payment_id:
            # Извлекаем db_id из уникального payment_id
            try:
                db_id = int(payment_id.split('_')[-1])
                c.execute("""
                    SELECT id, vk_id, personal_path, original_filename, state, status, disagree_reason, confirmed_at, created_at 
                    FROM vedomosti_users 
                    WHERE id = ? AND vk_id = ?
                """, (db_id, str(user_id)))
            except ValueError:
                # Если не удалось извлечь db_id, пробуем старый способ (поддерживаем и imported, и repet_imported)
                c.execute("""
                    SELECT id, vk_id, personal_path, original_filename, state, status, disagree_reason, confirmed_at, created_at 
                    FROM vedomosti_users 
                    WHERE vk_id = ?
                      AND (state = ? OR state = ?)
                """, (str(user_id), f"imported:{payment_id}", f"repet_imported:{payment_id}"))
        else:
            # Старый формат payment_id (поддерживаем и imported, и repet_imported)
            c.execute("""
                SELECT id, vk_id, personal_path, original_filename, state, status, disagree_reason, confirmed_at, created_at 
                FROM vedomosti_users 
                WHERE vk_id = ?
                  AND (state = ? OR state = ?)
                LIMIT 1
            """, (str(user_id), f"imported:{payment_id}", f"repet_imported:{payment_id}"))
        
        row = c.fetchone()
        conn.close()
        
        if not row:
            log.debug("Payment %s not found in DB for user %s", payment_id, user_id)
            return None
            
        log.debug("Found payment %s in DB for user %s", payment_id, user_id)
        db_id, vk_id_raw, personal_path, original_filename, state, status_db, disagree_reason_db, confirmed_at_db, created_at_db = row
        
        # Загружаем данные из CSV файла
        row_dict = {}
        if personal_path and os.path.exists(personal_path):
            try:
                df = get_cached_csv_data(personal_path)
                if isinstance(df, pd.DataFrame) and not df.empty:
                    row_dict = df.iloc[0].to_dict()
            except Exception:
                log.warning("Failed to read CSV for find_payment %s path=%s", payment_id, personal_path)
        
        # Определяем, является ли выплата репетиторской
        is_repet = False
        if state and state.startswith('repet_imported:'):
            is_repet = True
        
        # Маппинг данных в зависимости от типа выплаты
        if is_repet:
            payment_data = _map_row_to_repet_payment_data(row_dict, user_id, original_filename)
            payment_data['is_repet'] = True
        else:
            payment_data = _map_row_to_payment_data(row_dict, user_id, original_filename)
            payment_data['is_repet'] = False
        
        # ВАЖНО: Добавляем personal_path в данные для правильного отображения
        payment_data['personal_path'] = personal_path
        
        # Используем уникальный payment_id
        original_payment_id = state.split(':', 1)[1] if ':' in state else payment_id
        unique_payment_id = f"{original_payment_id}_{db_id}"
        
        entry = {
            "id": unique_payment_id,
            "data": payment_data,
            "created_at": float(created_at_db) if created_at_db else time.time(),
            "status": status_db or "new",
            "db_id": db_id,
            "original_payment_id": original_payment_id
        }
        
        if disagree_reason_db:
            entry["disagree_reason"] = disagree_reason_db
            
        log.info("Found payment %s for user %s in DB", payment_id, user_id)
        return entry
        
    except Exception:
        log.exception("Failed to find payment %s for user %s in DB", payment_id, user_id)
        return None

def send_payment_message(user_id: int, payment_entry: dict):
    """Отправляет сообщение с текстом выплаты и inline-кнопками."""
    is_repet = payment_entry.get("data", {}).get('is_repet', False)
    
    if is_repet:
        text = "У тебя появилась новая ведомость на согласование 📋\n\n" + format_repet_payment_text(payment_entry["data"])
    else:
        text = "У тебя появилась новая ведомость на согласование 📋\n\n" + format_payment_text(payment_entry["data"])
    
    keyboard = inline_confirm_keyboard(payment_id=payment_entry["id"])
    
    success = safe_vk_send(user_id, text, keyboard)
    if success:
        log.info("Отправлена выплата %s user=%s is_repet=%s", payment_entry["id"], user_id, is_repet)
    else:
        log.error("Не удалось отправить выплату %s user=%s", payment_entry.get("id"), user_id)


def simulate_two_payments_for_user(user_id: int, delay_seconds: int = 10):
    """Для ручного тестирования — не вызывается по умолчанию."""
    p1 = TEST_PAYMENT_BASE.copy()
    p1["vk_id"] = user_id
    p1["groups"] = "группа 1"
    pid1 = add_payment_for_user(user_id, p1)
    send_payment_message(user_id, find_payment(user_id, pid1))
    def send_second():
        p2 = TEST_PAYMENT_BASE.copy()
        p2["vk_id"] = user_id
        p2["groups"] = "группа 2"
        p2["fio"] = "Иван Иванов (вторая выплата)"
        pid2 = add_payment_for_user(user_id, p2)
        send_payment_message(user_id, find_payment(user_id, pid2))
        log.info("Вторая тестовая выплата отправлена user=%s", user_id)
    t = threading.Timer(delay_seconds, send_second)
    t.daemon = True
    t.start()


def handle_message_event(event):
    try:
        payload_str = None
        if hasattr(event, "object") and isinstance(event.object, dict):
            payload_str = event.object.get("payload")
            if not payload_str:
                msg = event.object.get("message") or {}
                payload_str = msg.get("payload")
        if not payload_str and hasattr(event, "raw"):
            payload_str = event.raw.get("object", {}).get("payload")
        payload = {}
        if payload_str:
            try:
                payload = json.loads(payload_str)
            except Exception:
                payload = {"raw": payload_str}
        else:
            payload = {}
        event_id = getattr(event, "event_id", None) or (event.object.get("event_id") if isinstance(event.object, dict) else None)
        user_id = getattr(event, "user_id", None) or (event.object.get("user_id") if isinstance(event.object, dict) else None)
        peer_id = getattr(event, "peer_id", None) or (event.object.get("peer_id") if isinstance(event.object, dict) else None)
        log.info("MESSAGE_EVENT payload=%s user=%s peer=%s event_id=%s", payload, user_id, peer_id, event_id)
        try:
            vk_session.method("messages.sendMessageEventAnswer", {
                "event_id": event_id,
                "user_id": user_id,
                "peer_id": peer_id,
                "payload": json.dumps({"type": "show_snackbar", "text": "Действие принято"}, ensure_ascii=False)
            })
        except Exception:
            log.debug("sendMessageEventAnswer failed", exc_info=True)
        cmd = payload.get("cmd")
        if cmd == "confirm_payment":
            choice = payload.get("choice")
            payment_id = payload.get("payment_id")
            p = find_payment(user_id, payment_id)
            is_repet = p.get("data", {}).get('is_repet', False) if p else False
            if choice == "agree":
                if p:
                    p["status"] = "agree_pending_verify"
                    data = p.get("data", {})
                    phone = data.get("phone", "-")
                    
                    # Преобразуем телефон в целое число (убираем .0 для кураторов)
                    if phone != "-":
                        try:
                            # Если телефон представлен как float (например, 79517249750.0)
                            if isinstance(phone, float):
                                phone = str(int(phone))
                            # Если телефон как строка с точкой
                            elif isinstance(phone, str) and phone.endswith('.0'):
                                phone = str(int(float(phone)))
                            # Если просто число как строка
                            elif phone.isdigit():
                                phone = str(int(phone))
                        except (ValueError, TypeError):
                            # Если что-то пошло не так, оставляем как есть
                            pass
                    
                    # Для репетиторов ФИО берём из 'fio', для кураторов — из 'console'
                    if is_repet:
                        console_name = data.get("fio") or data.get("console", "-")
                    else:
                        console_name = data.get("console", "-")
                    
                    text = (
                        "Проверь, пожалуйста, свои данные в приложении Консоль!\n\n"
                        f"Номер телефона получателя: {phone}\n"
                        f"ФИО получателя: {console_name}"
                    )
                safe_vk_send(user_id, text, yes_no_keyboard("agree_verify", payment_id))
                log.info("User %s started agree flow for payment %s (pending verify)", user_id, payment_id)
            else:
                if is_repet:
                    # Для репетиторов сразу фиксируем несогласие и логируем в отдельную таблицу
                    if p:
                        p["status"] = "disagreed"
                    try:
                        update_vedomosti_status_by_payment(payment_id, "disagreed", reason="Данные выплаты")
                    except Exception:
                        log.exception("Failed to persist disagree for repet payment %s", payment_id)
                    
                    filename = p.get("data", {}).get("original_filename", "") if p else ""
                    fio = p.get("data", {}).get("fio", "") if p else ""
                    log_repet_complaint_to_sheet(user_id, "Данные выплаты", filename, fio)
                    
                    safe_vk_send(user_id, "В сообщении ниже опишите причину несогласия. В течение n количества времени с Вами свяжется оператор.")
                    log.info("User %s disagreed repet payment %s -> logged to repet sheet", user_id, payment_id)
                else:
                    if p:
                        p["status"] = "disagree_select_point"
                    safe_vk_send(user_id, "С каким пунктом вы не согласны:", payments_disagree_keyboard(payment_id=payment_id))
                    log.info("User %s disagreed payment %s -> asking for point (no persist)", user_id, payment_id)
        elif cmd == "agree_verify":
            payment_id = payload.get("payment_id")
            choice = payload.get("choice")
            p = find_payment(user_id, payment_id)
            if not p:
                return
            if choice == "yes":
                p["status"] = "agree_pending_pro"
                safe_vk_send(user_id, "Подписан ли у Вас договор в приложении Консоль Про? Краткая справка, как проверить: Консоль Про ->  раздел компании. Если там есть компания ООО '100балльный репетитор', то договор подписан.", yes_no_keyboard("agree_pro", payment_id))
                log.info("User %s verified data for payment %s", user_id, payment_id)
            else:
                p["status"] = "agree_data_mismatch"
                safe_vk_send(user_id, "С Вами свяжется оператор. Пожалуйста, напишите в сообщении ниже корректные данные.")
                log.info("User %s reported data mismatch for payment %s", user_id, payment_id)
                p = find_payment(user_id, payment_id)
                filename = p.get("data", {}).get("original_filename", "") if p else ""
                filepath = f"hosting/open/{filename}" if filename else ""
                fio = ""
                is_repet = False
                if p:
                    data = p.get("data", {}) or {}
                    fio = data.get("fio") or data.get("curator") or ""
                    is_repet = data.get("is_repet", False)
                if is_repet:
                    log_repet_complaint_to_sheet(user_id, "Не те данные в Консоли", filename, fio)
                else:
                    log_complaint_to_sheet(user_id, "Не те данные в Консоли", filename, filepath, fio)
        elif cmd == "agree_pro":
            payment_id = payload.get("payment_id")
            choice = payload.get("choice")
            p = find_payment(user_id, payment_id)
            if not p:
                return
            if choice == "yes":
                p["status"] = "agreed"
                try:
                    update_vedomosti_status_by_payment(payment_id, "agreed")
                except Exception:
                    log.exception("Failed to persist agree status for payment %s", payment_id)
                safe_vk_send(user_id, " Вы согласовали выплату. Спасибо! В течение 10 дней в приложении Консоль Вам придет акт, который необходимо подписать. После этого в течение n количества времени на реквизиты Вашего банковского счета придет выплата.")
                log.info("User %s agreed payment %s after PRO confirmation", user_id, payment_id)
            else:
                p["status"] = "agree_pro_pending"
                safe_vk_send(user_id, "Прими приглашение в Консоль ПРО, затем повторно подтвердите выплату.")
                # Логируем отказ принять приглашение в Консоль ПРО в таблицу, чтобы оператор увидел
                filename = p.get("data", {}).get("original_filename", "") if p else ""
                filepath = f"hosting/open/{filename}" if filename else ""
                fio = ""
                is_repet = False
                if p:
                    data = p.get("data", {}) or {}
                    fio = data.get("fio") or data.get("curator") or ""
                    is_repet = data.get("is_repet", False)
                if is_repet:
                    log_repet_complaint_to_sheet(user_id, "Не принял приглашение в Консоль ПРО", filename, fio)
                else:
                    log_complaint_to_sheet(user_id, "Не принял приглашение в Консоль ПРО", filename, filepath, fio)
                log.info("User %s has not accepted PRO invite for payment %s", user_id, payment_id)
        elif cmd == "disagree_reason":
            sid = payload.get("payment_id")
            reason = payload.get("reason")
            p = find_payment(user_id, sid)
            if reason == "Иная причина (связаться с оператором)":
                if p:
                    p["status"] = "disagreed"
                try:
                    update_vedomosti_status_by_payment(sid, "disagreed", reason=reason)
                except Exception:
                    log.exception("Failed to persist disagree (other reason) for %s", sid)
                filename = p.get("data", {}).get("original_filename", "") if p else ""
                filepath = f"hosting/open/{filename}" if filename else ""
                fio = ""
                is_repet = False
                if p:
                    data = p.get("data", {}) or {}
                    fio = data.get("fio") or data.get("curator") or ""
                    is_repet = data.get("is_repet", False)
                if is_repet:
                    log_repet_complaint_to_sheet(user_id, "Иная причина (связаться с оператором)", filename, fio)
                else:
                    log_complaint_to_sheet(user_id, "Иная причина (связаться с оператором)", filename, filepath, fio)
                safe_vk_send(user_id, "Сообщение передано оператору. Он скоро свяжется с Вами. Опишите, пожалуйста, Вашу проблему в сообщении ниже.")
                log.info("User %s chose other reason for %s -> operator handoff", user_id, sid)
                return
            if p:
                p["disagree_reason"] = reason
            filename = p.get("data", {}).get("original_filename", "") if p else ""
            filepath = f"hosting/open/{filename}" if filename else ""
            is_repet = False
            if p:
                is_repet = (p.get("data", {}) or {}).get("is_repet", False)
            if is_repet:
                log_repet_complaint_to_sheet(user_id, f"Выбран пункт несогласия: {reason}", filename, (p.get("data", {}) or {}).get("fio", ""))
            else:
                log_complaint_to_sheet(user_id, f"Выбран пункт несогласия: {reason}", filename, filepath)
            conflict_type = map_reason_to_type(reason)
            if conflict_type:
                file_base = None
                personal_path = None
                try:
                    payment_data = (p or {}).get("data", {})
                    fname = payment_data.get("original_filename")
                    if fname:
                        file_base = os.path.splitext(fname)[0]
                    personal_path = payment_data.get("personal_path")
                except Exception:
                    file_base = None
                    personal_path = None
                try:
                    explanation = format_conflict(file_base or "", user_id, conflict_type, personal_path=personal_path)
                    if explanation:
                        safe_vk_send(user_id, explanation, disagreement_decision_keyboard(payment_id=sid, reason_label=reason))
                except Exception:
                    log.debug("Failed to send conflict explanation", exc_info=True)
            log.info("User %s set disagree reason for payment %s -> %s (awaiting decision)", user_id, sid, reason)
        elif cmd == "disagree_decision":
            sid = payload.get("payment_id")
            choice = payload.get("choice")
            reason = payload.get("reason")
            p = find_payment(user_id, sid)
            if not p:
                return
            if choice == "agree_point":
                # Переадресация на общий список пунктов
                safe_vk_send(user_id, "С каким пунктом вы не согласны:", payments_disagree_keyboard(payment_id=sid))
                log.info("User %s decided agree_point for %s -> redirect to general list", user_id, sid)
            elif choice == "disagree_point":
                p["status"] = "disagreed"
                try:
                    update_vedomosti_status_by_payment(sid, "disagreed", reason=reason)
                except Exception:
                    log.exception("Failed to persist disagreed for %s", sid)
                filename = p.get("data", {}).get("original_filename", "") if p else ""
                filepath = f"hosting/open/{filename}" if filename else ""
                data = p.get("data", {}) or {}
                is_repet = data.get("is_repet", False)
                if is_repet:
                    log_repet_complaint_to_sheet(user_id, f"Не согласен с пунктом: {reason}", filename, data.get("fio", ""))
                else:
                    log_complaint_to_sheet(user_id, f"Не согласен с пунктом: {reason}", filename, filepath)
                safe_vk_send(user_id, "Сообщение передано оператору. Он скоро свяжется с Вами. Опишите, пожалуйста, Вашу проблему в сообщении ниже.")
                log.info("User %s decided disagree_point for %s (persisted)", user_id, sid)
        elif cmd == "agree_payment":
            # Обработка кнопки "Согласиться с ведомостью" из общего списка
            sid = payload.get("payment_id")
            p = find_payment(user_id, sid)
            if not p:
                return
            # Промежуточное сообщение с подтверждением
            safe_vk_send(user_id, "Вы точно согласны с ведомостью?", final_agreement_keyboard(sid))
            log.info("User %s clicked agree payment %s -> showing final agreement", user_id, sid)
        elif cmd == "final_agreement":
            # Обработка финального подтверждения согласия
            sid = payload.get("payment_id")
            choice = payload.get("choice")
            p = find_payment(user_id, sid)
            if not p:
                return
            if choice == "yes":
                p["status"] = "agree_pending_verify"
                data = p.get("data", {})
                phone = data.get("phone", "-")
                is_repet = data.get('is_repet', False)
                
                # Преобразуем телефон, убирая .0 если есть
                if phone != "-":
                    # Если телефон приходит как float (например, 79517249750.0)
                    if isinstance(phone, float):
                        phone = str(int(phone))
                    # Если телефон как строка с .0
                    elif isinstance(phone, str) and phone.endswith('.0'):
                        phone = str(int(float(phone)))
                    # Если просто число как строка (без .0)
                    elif isinstance(phone, str) and phone.replace('.', '', 1).isdigit():
                        # Проверяем, есть ли точка
                        if '.' in phone:
                            phone = str(int(float(phone)))
                        else:
                            phone = str(int(phone))
                
                # Для репетиторов ФИО берём из 'fio', для кураторов — из 'console'
                if is_repet:
                    console_name = data.get("fio") or data.get("console", "-")
                else:
                    console_name = data.get("console", "-")
                
                text = (
                    "Проверь, пожалуйста, свои данные в приложении Консоль!\n\n"
                    f"Номер телефона получателя: {phone}\n"
                    f"ФИО получателя: {console_name}"
                )
                safe_vk_send(user_id, text, yes_no_keyboard("agree_verify", sid))
                log.info("User %s confirmed final agreement for payment %s", user_id, sid)
            else:
                # Переадресация на общий список пунктов
                safe_vk_send(user_id, "С каким пунктом вы не согласны:", payments_disagree_keyboard(payment_id=sid))
                log.info("User %s declined final agreement for payment %s -> redirect to general list", user_id, sid)
        elif cmd == "open_statement":
            sid = payload.get("statement_id")
            p = find_payment(user_id, sid)
            if p:
                log.info("User %s trying to open statement %s with status: %s", user_id, sid, p.get("status"))
                current_status = refresh_payment_status_from_db(p, user_id)
                if current_status == "agreed":
                    safe_vk_send(user_id, "Вы уже согласовали ведомость!")
                    log.info("User %s tried to open already confirmed statement %s", user_id, sid)
                    return
                
                # Выбираем форматирование в зависимости от типа выплаты (куратор / репетитор)
                if p.get("data", {}).get("is_repet"):
                    body = format_repet_payment_text(p["data"])
                else:
                    body = format_payment_text(p["data"])
                statement_text = body
                safe_vk_send(user_id, statement_text, inline_confirm_keyboard(payment_id=sid))
                user_last_opened_payment[user_id] = sid  # Запоминаем последнюю открытую выплату
                log.info("User %s opened statement %s (unique_payment_id=%s)", user_id, sid, sid)
            else:
                safe_vk_send(user_id, "Ведомость не найдена (возможно устарела).")
        elif cmd == "to_list":
            payments = get_all_payments_for_user_from_db(user_id)
            send_payments_list_multiple(user_id, payments, page=0)
            log.info("Sent payments list to %s", user_id)
            return
        elif cmd == "payments_page":
            page = int(payload.get("page", 0))
            payments = get_all_payments_for_user_from_db(user_id)
            send_payments_list_multiple(user_id, payments, page=page)
            return
        else:
            safe_vk_send(user_id, f"Нажата inline-кнопка. Payload: {json.dumps(payload, ensure_ascii=False)}")
    except Exception:
        log.exception("Ошибка в handle_message_event: %s", traceback.format_exc())


def handle_message_new(event):
    try:
        msg = event.object.get("message") if isinstance(event.object, dict) else getattr(event, "message", None)
        if not msg:
            return
        text = (msg.get("text") or "").strip()
        payload_str = msg.get("payload") or msg.get("data")
        from_id = msg.get("from_id") or msg.get("peer_id")
        peer_id = msg.get("peer_id") or from_id
        log.info("MESSAGE_NEW from=%s text=%s payload=%s", from_id, text, bool(payload_str))
        if text in ("Согласен с выплатой", "Не согласен с выплатой"):
            # Используем последнюю открытую выплату
            last_payment_id = user_last_opened_payment.get(from_id)
            log.info("User %s using button agreement, last_payment_id=%s", from_id, last_payment_id)
            if not last_payment_id:
                safe_vk_send(from_id, "Сначала откройте ведомость из списка выплат.")
                return
            p = find_payment(from_id, last_payment_id)
            log.info("User %s find_payment result for %s: %s", from_id, last_payment_id, "Found" if p else "Not found")
            if p:
                log.info("Found payment details: db_id=%s, original_filename=%s, status=%s", 
                        p.get("db_id"), p.get("data", {}).get("original_filename"), p.get("status"))
            if not p:
                safe_vk_send(from_id, "Ведомость не найдена. Откройте ведомость заново.")
                return
            pid = p["id"]
            is_repet = p.get("data", {}).get('is_repet', False)
            if text == "Согласен с выплатой":
                p["status"] = "agree_pending_verify"
                data = p.get("data", {})
                phone = data.get("phone", "-")
                if phone != "-":
                    # Если телефон приходит как float (например, 79517249750.0)
                    if isinstance(phone, float):
                        phone = str(int(phone))
                    # Если телефон как строка с .0
                    elif isinstance(phone, str) and phone.endswith('.0'):
                        phone = str(int(float(phone)))
                    # Если просто число как строка (без .0)
                    elif isinstance(phone, str) and phone.replace('.', '', 1).isdigit():
                        # Проверяем, есть ли точка
                        if '.' in phone:
                            phone = str(int(float(phone)))
                        else:
                            phone = str(int(phone))
                # Для репетиторов ФИО берём из 'fio', для кураторов — из 'console'
                if is_repet:
                    console_name = data.get("fio") or data.get("console", "-")
                else:
                    console_name = data.get("console", "-")
                text_msg = (
                    "[Проверь, пожалуйста, свои данные в приложении Консоль!]\n\n"
                    f"Номер телефона получателя: {phone}\n"
                    f"ФИО получателя: {console_name}"
                )
                safe_vk_send(from_id, text_msg, yes_no_keyboard("agree_verify", pid))
                log.info("User %s started agree flow via text-button for %s", from_id, pid)
                return
            else:
                if is_repet:
                    # Для репетиторов сразу фиксируем несогласие и логируем в отдельную таблицу
                    p["status"] = "disagreed"
                    try:
                        update_vedomosti_status_by_payment(pid, "disagreed", reason="Данные выплаты")
                    except Exception:
                        log.exception("Failed to persist disagree for repet payment %s", pid)
                    
                    filename = p.get("data", {}).get("original_filename", "")
                    fio = p.get("data", {}).get("fio", "")
                    log_repet_complaint_to_sheet(from_id, "Данные выплаты", filename, fio)
                    
                    safe_vk_send(from_id, "В сообщении ниже опишите причину несогласия. В течение n количества времени с Вами свяжется оператор.")
                    log.info("User %s disagreed repet payment %s via text-button -> logged to repet sheet", from_id, pid)
                else:
                    p["status"] = "disagree_select_point"
                    safe_vk_send(from_id, "С каким пунктом вы не согласны:", payments_disagree_keyboard(payment_id=pid))
                    log.info("User %s disagreed payment %s via text-button -> asking for point (no persist)", from_id, pid)
                    return
        if payload_str:
            try:
                payload = json.loads(payload_str)
            except Exception:
                payload = {"raw": payload_str}
            cmd = payload.get("cmd")
            if cmd == "open_statement":
                sid = payload.get("statement_id")
                p = find_payment(from_id, sid)
                if p:
                    log.info("User %s trying to open statement %s via payload with status: %s", from_id, sid, p.get("status"))
                    current_status = refresh_payment_status_from_db(p, from_id)
                    if current_status == "agreed":
                        vk.messages.send(
                            user_id=from_id,
                            random_id=vk_api.utils.get_random_id(),
                            message="Вы уже согласовали ведомость!"
                        )
                        log.info("User %s tried to open already confirmed statement %s via payload", from_id, sid)
                        return
                    
                    if p.get("data", {}).get("is_repet"):
                        body = format_repet_payment_text(p["data"])
                    else:
                        body = format_payment_text(p["data"])
                    statement_text = body
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message=statement_text,
                        keyboard=inline_confirm_keyboard(payment_id=sid)
                    )
                    user_last_opened_payment[from_id] = sid  # Запоминаем последнюю открытую выплату
                    log.info("User %s opened statement %s via payload (unique_payment_id=%s)", from_id, sid, sid)
                    return
                else:
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="Ведомость не найдена (возможно устарела)."
                    )
                    return
            if cmd == "disagree_reason":
                sid = payload.get("payment_id")
                reason = payload.get("reason")
                p = find_payment(from_id, sid)
                if reason == "Иная причина (связаться с оператором)":
                    if p:
                        p["status"] = "disagreed"
                    try:
                        update_vedomosti_status_by_payment(sid, "disagreed", reason=reason)
                    except Exception:
                        log.exception("Failed to persist disagree (other reason) for %s", sid)
                    filename = p.get("data", {}).get("original_filename", "") if p else ""
                    filepath = f"hosting/open/{filename}" if filename else ""
                    data = p.get("data", {}) or {}
                    is_repet = data.get("is_repet", False)
                    if is_repet:
                        log_repet_complaint_to_sheet(from_id, "Иная причина (связаться с оператором)", filename, data.get("fio", ""))
                    else:
                        log_complaint_to_sheet(from_id, "Иная причина (связаться с оператором)", filename, filepath)
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="Сообщение передано оператору. Он скоро свяжется с Вами. Опишите, пожалуйста, Вашу проблему в сообщении ниже."
                    )
                    return
                if p:
                    p["disagree_reason"] = reason
                conflict_type = map_reason_to_type(reason)
                if conflict_type:
                    file_base = None
                    personal_path = None
                    try:
                        payment_data = (p or {}).get("data", {})
                        fname = payment_data.get("original_filename")
                        if fname:
                            file_base = os.path.splitext(fname)[0]
                        personal_path = payment_data.get("personal_path")
                    except Exception:
                        file_base = None
                        personal_path = None
                    try:
                        explanation = format_conflict(file_base or "", from_id, conflict_type, personal_path=personal_path)
                        if explanation:
                            vk.messages.send(
                                user_id=from_id,
                                random_id=vk_api.utils.get_random_id(),
                                message=explanation,
                                keyboard=disagreement_decision_keyboard(payment_id=sid, reason_label=reason)
                            )
                    except Exception:
                        log.debug("Failed to send conflict explanation (MESSAGE_NEW)", exc_info=True)
                log.info("User %s set disagree reason via MESSAGE_NEW for payment %s -> %s (awaiting decision)", from_id, sid, reason)
                return
            if cmd == "disagree_decision":
                sid = payload.get("payment_id")
                choice = payload.get("choice")
                reason = payload.get("reason")
                p = find_payment(from_id, sid)
                if not p:
                    return
                if choice == "agree_point":
                    # Переадресация на общий список пунктов
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="С каким пунктом вы не согласны:",
                        keyboard=payments_disagree_keyboard(payment_id=sid)
                    )
                elif choice == "disagree_point":
                    p["status"] = "disagreed"
                    try:
                        update_vedomosti_status_by_payment(sid, "disagreed", reason=reason)
                    except Exception:
                        log.exception("Failed to persist disagreed for %s", sid)
                    filename = p.get("data", {}).get("original_filename", "") if p else ""
                    filepath = f"hosting/open/{filename}" if filename else ""
                    data = p.get("data", {}) or {}
                    is_repet = data.get("is_repet", False)
                    if is_repet:
                        log_repet_complaint_to_sheet(from_id, f"Не согласен с пунктом: {reason}", filename, data.get("fio", ""))
                    else:
                        log_complaint_to_sheet(from_id, f"Не согласен с пунктом: {reason}", filename, filepath)
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="Сообщение передано оператору. Он скоро свяжется с Вами. Опишите, пожалуйста, Вашу проблему в сообщении ниже."
                    )
                return
            if cmd == "agree_payment":
                # Обработка кнопки "Согласиться с ведомостью" из общего списка
                sid = payload.get("payment_id")
                p = find_payment(from_id, sid)
                if not p:
                    return
                # Промежуточное сообщение с подтверждением
                vk.messages.send(
                    user_id=from_id,
                    random_id=vk_api.utils.get_random_id(),
                    message="Вы точно согласны с ведомостью?",
                    keyboard=final_agreement_keyboard(sid)
                )
                return
            if cmd == "final_agreement":
                # Обработка финального подтверждения согласия
                sid = payload.get("payment_id")
                choice = payload.get("choice")
                p = find_payment(from_id, sid)
                if not p:
                    return
                if choice == "yes":
                    p["status"] = "agree_pending_verify"
                    data = p.get("data", {})
                    phone = data.get("phone", "-")
                    is_repet = data.get('is_repet', False)
                    
                    # Преобразуем телефон, убирая .0 если есть
                    if phone != "-":
                        # Если телефон приходит как float (например, 79517249750.0)
                        if isinstance(phone, float):
                            phone = str(int(phone))
                        # Если телефон как строка с .0
                        elif isinstance(phone, str) and phone.endswith('.0'):
                            phone = str(int(float(phone)))
                        # Если просто число как строка (без .0)
                        elif isinstance(phone, str) and phone.replace('.', '', 1).isdigit():
                            # Проверяем, есть ли точка
                            if '.' in phone:
                                phone = str(int(float(phone)))
                            else:
                                phone = str(int(phone))
                    
                    # Для репетиторов ФИО берём из 'fio', для кураторов — из 'console'
                    if is_repet:
                        console_name = data.get("fio") or data.get("console", "-")
                    else:
                        console_name = data.get("console", "-")
                    
                    text = (
                        "Проверь, пожалуйста, свои данные в приложении Консоль!\n\n"
                        f"Номер телефона получателя: {phone}\n"
                        f"ФИО получателя: {console_name}"
                    )
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message=text,
                        keyboard=yes_no_keyboard("agree_verify", sid)
                    )
                else:
                    # Переадресация на общий список пунктов
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="С каким пунктом вы не согласны:",
                        keyboard=payments_disagree_keyboard(payment_id=sid)
                    )
                return
            if cmd == "to_list":
                payments = get_all_payments_for_user_from_db(from_id)
                if not payments:
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="У Вас нет выплат.",
                        keyboard=chat_bottom_keyboard()
                    )
                    return
                statements = []
                for idx, p in enumerate(payments, start=1):
                    label = _format_payment_label(
                        p["data"].get('original_filename'), 
                        idx,
                        30,
                        p.get("created_at"),
                        p.get("db_id"),
                        p.get("data", {}).get("groups")  # Передаем информацию о группах
                    )
                    statements.append((p["id"], label))
                send_payments_list_multiple(from_id, payments, page=0, use_vk_direct=True)
                log.info("Sent payments list to %s", from_id)
                return
            if cmd == "payments_page":
                page = int(payload.get("page", 0))
                payments = get_all_payments_for_user_from_db(from_id)
                send_payments_list_multiple(from_id, payments, page=page, use_vk_direct=True)
                return
            if cmd == "agree_verify":
                sid = payload.get("payment_id")
                choice = payload.get("choice")
                p = find_payment(from_id, sid)
                if not p:
                    return
                if choice == "yes":
                    p["status"] = "agree_pending_pro"
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="Подписан ли у Вас договор в приложении Консоль Про? Краткая справка, как проверить: Консоль Про ->  раздел компании. Если там есть компания ООО '100балльный репетитор', то договор подписан. ",
                        keyboard=yes_no_keyboard("agree_pro", sid)
                    )
                else:
                    p["status"] = "agree_data_mismatch"
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="С Вами свяжется оператор."
                    )
                    filename = p.get("data", {}).get("original_filename", "") if p else ""
                    filepath = f"hosting/open/{filename}" if filename else ""
                    data = p.get("data", {}) or {}
                    is_repet = data.get("is_repet", False)
                    if is_repet:
                        log_repet_complaint_to_sheet(from_id, "Не те данные в Консоли", filename, data.get("fio", ""))
                    else:
                        log_complaint_to_sheet(from_id, "Не те данные в Консоли", filename, filepath)
                return
            if cmd == "agree_pro":
                sid = payload.get("payment_id")
                choice = payload.get("choice")
                p = find_payment(from_id, sid)
                if not p:
                    return
                if choice == "yes":
                    p["status"] = "agreed"
                    try:
                        update_vedomosti_status_by_payment(sid, "agreed")
                    except Exception:
                        log.exception("Failed to persist agree status for payment %s", sid)
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message=" Вы согласовали выплату. Спасибо! В течение 10 дней в приложении Консоль Вам придет акт, который необходимо подписать. После этого в течение n количества времени на реквизиты Вашего банковского счета придет выплата."
                    )
                else:
                    p["status"] = "agree_pro_pending"
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="С Вами свяжется оператор."
                    )
                    # Логируем отказ принять приглашение в Консоль ПРО в таблицу
                    filename = p.get("data", {}).get("original_filename", "") if p else ""
                    filepath = f"hosting/open/{filename}" if filename else ""
                    data = p.get("data", {}) or {}
                    is_repet = data.get("is_repet", False)
                    if is_repet:
                        log_repet_complaint_to_sheet(from_id, "Не подписал договор", filename, data.get("fio", ""))
                    else:
                        log_complaint_to_sheet(from_id, "Не подписал договор", filename, filepath)
                return
        if text.lower() == "к списку выплат" or text == "К списку выплат":
            payments = get_all_payments_for_user_from_db(from_id)
            if not payments:
                vk.messages.send(
                    user_id=from_id,
                    random_id=vk_api.utils.get_random_id(),
                    message="У Вас нет выплат.",
                    keyboard=chat_bottom_keyboard()
                )
                return
            statements = []
            for idx, p in enumerate(payments, start=1):
                label = _format_payment_label(p["data"].get('original_filename'), idx)
                statements.append((p["id"], label))
            send_payments_list_multiple(from_id, payments, page=0, use_vk_direct=True)
            log.info("Sent payments list to %s", from_id)
            return
        m = re.match(r"^\s*Ведомость\s+(\d+)\s*$", text, flags=re.IGNORECASE)
        if m:
            idx = int(m.group(1)) - 1
            payments = get_all_payments_for_user_from_db(from_id)
            if 0 <= idx < len(payments):
                p = payments[idx]
                log.info("User %s trying to open statement %s by text with status: %s", from_id, p["id"], p.get("status"))
                current_status = refresh_payment_status_from_db(p, from_id)
                if current_status == "agreed":
                    vk.messages.send(
                        user_id=from_id,
                        random_id=vk_api.utils.get_random_id(),
                        message="Вы уже согласовали ведомость!"
                    )
                    log.info("User %s tried to open already confirmed statement %s by text", from_id, p["id"])
                    return
                if p.get("data", {}).get("is_repet"):
                    body = format_repet_payment_text(p["data"])
                else:
                    body = format_payment_text(p["data"])
                statement_text = body
                vk.messages.send(
                    user_id=from_id,
                    random_id=vk_api.utils.get_random_id(),
                    message=statement_text,
                    keyboard=inline_confirm_keyboard(payment_id=p["id"]) )
                user_last_opened_payment[from_id] = p["id"]  # Запоминаем последнюю открытую выплату
                log.info("User %s opened statement %s by text click", from_id, p["id"]) 
                return
            else:
                vk.messages.send(
                    user_id=from_id,
                    random_id=vk_api.utils.get_random_id(),
                    message="Ведомость не найдена (неверный номер)."
                )
                return
        # Приветственное сообщение показываем только при явной команде "Начать"/"/start"
        if text.lower() in ("начать", "/start", "start"):
            vk.messages.send(
                user_id=from_id,
                random_id=vk_api.utils.get_random_id(),
                message="Я бот по согласованию выплат. Нажмите кнопку 'К списку выплат' или дождитесь уведомления о выплате.",
                keyboard=chat_bottom_keyboard()
            )
    except Exception:
        log.exception("Ошибка в handle_message_new: %s", traceback.format_exc())

def main_loop():
    log.info("Бот запущен. Ожидание событий...")
    ensure_vedomosti_status_columns()
    ensure_unique_import_states()
    ensure_db_indexes()  # Создаем индексы для оптимизации
    try:
        loaded = load_imported_vedomosti_into_memory(send_notifications=False)
        log.info("Startup: loaded %d existing imported vedomosti into memory", loaded)
    except Exception:
        log.exception("Failed during startup loading of imported vedomosti")
    importer_thread = threading.Thread(target=background_importer, args=(5.0,), daemon=True)
    importer_thread.start()
    for event in longpoll.listen():
        try:
            if event.type == VkBotEventType.MESSAGE_EVENT:
                handle_message_event(event)
            elif event.type == VkBotEventType.MESSAGE_NEW:
                handle_message_new(event)
            else:
                pass
        except Exception:
            log.exception("Ошибка в основном loop: %s", traceback.format_exc())
        time.sleep(0.05)

if __name__ == "__main__":
    try:
        main_loop()
    except KeyboardInterrupt:
        log.info("Выключение по Ctrl+C")
    except Exception:
        log.exception("Критическая ошибка")
